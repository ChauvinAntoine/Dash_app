# -*- coding: utf-8 -*-

from dash import Dash, html, dcc, dash_table
from dash.dependencies import Input, Output, State

import menu
import section

#----------------------------------------------------------------------------------
#                        Accès à l'application
#-----------------------------------------------------------------------------------
# link : http://127.0.0.1:8050


#----------------------------------------------------------------------------------
#                        A MODIFIER
#-----------------------------------------------------------------------------------
# please modify with your personal path
#path = "C:/Users/achauvin/Business et Decision/Orange - Data & IA/Dash_v2 - ops"




#----------------------------------------------------------------------------------
#                        PACKAGES IMPORT
#-----------------------------------------------------------------------------------
import os
import base64
import io
import datetime
import itertools
import math
import openpyxl
#You may have to install Dash package
#import dash_table
#import dash_core_components as dcc
import pandas as pd
from sklearn.cluster import KMeans, DBSCAN
#import matplotlib.pyplot as plt
import numpy as np
#You may have to install ADTK package
from adtk.detector import PersistAD, VolatilityShiftAD
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from datetime import timedelta
from sklearn.neighbors import NearestNeighbors
from kneebow.rotor import Rotor
from operator import itemgetter
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.graph_objects as go

from scipy import stats

#################### Création de l'application ####################
app = Dash(
    __name__,
    assets_folder= 'assets/',
    suppress_callback_exceptions=True
)


# image_filename_logo = path+'\\Logo.jpg'
# encoded_image_logo = base64.b64encode(open(image_filename_logo, 'rb').read()).decode('ascii')

# image_filename_ds = path+'\\data_science.png'
# encoded_image_ds = base64.b64encode(open(image_filename_ds, 'rb').read()).decode('ascii')

# image_filename_orange = path+'\\Orange.png'
# encoded_image_orange = base64.b64encode(open(image_filename_orange, 'rb').read()).decode('ascii')

#################### Design de l'application ####################
app.layout = html.Div(
    className="body_app",
    children=[

        # Menu du dashboard
        menu.lef_menu(),

        # Contenu du dashboard
        html.Div(
            className="right_content",
            children=[
                html.Header([html.H1("Interference detection", className="titre_dashboard")], className="block_titre_dashboard"),
                #section.section_00(),
                section.section_01(),
                section.section_0(),
                section.section_1(),
                section.section_2(),
                section.section_3(),
                #section.section_6(),
                section.section_4(),
                #section.section_5(),
            ]
        )
    ])



#################### Elements interactifs ####################
@app.callback(Output('detection_course_status_start', 'children'),
 			[Input('graph_button_0','n_clicks')]) #,
             #Input('dropdown_sort','value')])

def detection_course_status(n_clicks) : #, prioritization) :
    if n_clicks >= 1 :
        return(html.H6('Detection course is running, please wait.'))
    else :
        return(None)

@app.callback([Output('graph_preclust', 'children'),
               Output('table_preclust', 'children'),
               Output('detailled_output_table','children'),
               #Output('table_ext_interf', 'children'),
               #Output('dropdown_interf', 'children'),
               #Output('loc_table','children'),
               Output('detection_course_status_end', 'children'),
               Output('export_excel_button', 'children')],
 			[Input('graph_button_0','n_clicks'),
              #Input('select_file','value'),
              Input('input-max_rtwp','value'),
              Input('input-miss_data','value'),
               #Input('input-k_pc','value'),
               Input('dropdown_detection_level','value'),
               #Input('slider_kmeans','value'),
               #Input('slider_riskobs','value'),
              Input('checklist_qos','value'),
              Input('checklist_prio','value'),
             Input('input-DCR-PS','value'),
             Input('input-CSSR-CS','value'),
             Input('input-CSSR-PS','value')]) #,
             #Input('dropdown_sort','value')])

def detection_course(n_clicks, max_RTWP, missing_data, detection_level, qos_kpi, indicators, poids_DCRPS, poids_CSSRCS, poids_CSSRPS) : #, prioritization) :
    if n_clicks >= 1 :
        try :
            for filename in os.listdir('data_dash'):
                # print(filename)
                if filename not in (['df_total.csv', 'df_loc.csv', 'crown_data.csv']):
                    os.remove("data_dash/" + filename)

            k_pc = 10
            k_clust = 300
            riskobs = 0

            # if data == 'Abidjan' :
            #     df = pd.read_csv('df - Abidjan.csv')
            #     df_loc = pd.read_csv('df_loc - Abidjan.csv')
            #     crown_data = pd.read_csv('crown_data_Abidjan.csv')
            # elif data == 'Senegal' :
            #     df = pd.read_csv('df - Senegal.csv')
            #     df_loc = pd.read_csv('df_loc - Senegal.csv')
            #     crown_data = pd.read_csv('crown_data_Senegal.csv')
            # elif data == 'Caribbean' :
            #     df = pd.read_csv('df - Caraibes.csv')
            #     df_loc = pd.read_csv('df_loc - Caraibes.csv')

            df = pd.read_csv('data_dash/df_total.csv')
            df_loc = pd.read_csv('data_dash/df_loc.csv')
            crown_data = pd.read_csv('data_dash/crown_data.csv')

            #df.to_csv('data_dash/df_total.csv')
            #crown_data.to_csv('data_dash/crown_data.csv')

            df_total = df
            df = df.dropna()
            df_loc = df_loc[['Cell','Latitude','Longitude']]
            df_loc = df_loc.drop_duplicates()
            df_loc.to_csv('data_dash/df_loc.csv')

            for i in df.columns :
                if i != 'Time' and i != 'Cell' :
                    df.loc[:,i] = df.loc[:,i].astype(float)

            df.replace('None', pd.np.nan, inplace=True)
            df.dropna(axis=0, inplace=True)

            if '3G DCR ALL DATA SERVICES' in df.columns :
                index_drop = df[(df['3G DCR ALL DATA SERVICES'] > 95)].index
                df.drop(index_drop, inplace = True)

            if 'DCR-PS' in df.columns :
                index_drop = df[(df['DCR-PS'] > 95)].index
                df.drop(index_drop, inplace = True)

            if 'RRC Succ Ratio' in df.columns :
                index_drop = df[(df['RRC Succ Ratio'] < 0.1)].index
                df.drop(index_drop, inplace = True)

            if 'CSSR-CS' in df.columns :
                index_drop = df[(df['CSSR-CS'] < 5)].index
                df.drop(index_drop, inplace = True)

            if 'CSSR-PS' in df.columns :
                index_drop = df[(df['CSSR-PS'] < 5)].index
                df.drop(index_drop, inplace = True)

            for i in range(len(df.columns)):
                if df.iloc[:,i].name not in ['3G DCR ALL DATA SERVICES', 'DCR-PS', 'CSSR-CS', 'CSSR-PS', 'RRC Succ Ratio', 'Cell', 'Time']:
                    Q75 = df.iloc[:,i].quantile(0.75)
                    Q25 = df.iloc[:,i].quantile(0.25)
                    cutoff = 8*(Q75 - Q25)
                    index_drop = df[(df.iloc[:,i] > Q75 + cutoff) | (df.iloc[:,i] < Q25 - cutoff)].index
                    df.drop(index_drop, inplace=True)

            df_data = df.drop_duplicates()

            df_data = df_data.merge(df_loc, left_on = 'Cell', right_on = 'Cell')
            #df_loc = df_data[['Cell','RTWP','Avg Num HSUPA Users','Latitude','Longitude']].groupby('Cell').agg('mean')

            missing_cells = []
            for cell in list(df.Cell.unique()) :
                if cell not in list(df_data.Cell.unique()) :
                    missing_cells.append(cell)

            for cell in df_data.Cell.unique() :
                df_total_cell = df_total[df_total.Cell == cell]
                df_data_cell = df_data[df_data.Cell == cell]
                if len(df_data_cell)/len(df_total_cell) < (100-missing_data)/100 :
                    missing_cells.append(cell)

            new_df_data = pd.DataFrame()
            for cell in df_data.Cell.unique() :
                if cell not in list(missing_cells) :
                    df_cell = df_data[df_data.Cell == cell]
                    new_df_data = pd.concat([new_df_data,df_cell])

            df_data = new_df_data

            df_clean = df_data
            if 'DCR-PS' not in df_clean.columns and '3G DCR ALL DATA SERVICES' in df_clean.columns:
                df_clean['DCR-PS'] = df_clean['3G DCR ALL DATA SERVICES']
            df_clean.to_csv('data_dash/df_clean.csv')

            risk_cell_dbm = [] ; new_df_data = pd.DataFrame()
            for cell in df_data.Cell.unique() :
                df_cell = df_data[df_data.Cell == cell]
                if max(df_cell.RTWP) < max_RTWP :
                    new_df_data = pd.concat([new_df_data,df_cell])
                else :
                    risk_cell_dbm.append(cell)

            risk_cell_dbm = pd.DataFrame(data = {'Cell' : risk_cell_dbm})

            df_data = new_df_data
            df_data.to_csv('data_dash/df_data.csv')

            missing_cells = pd.DataFrame(data = {'Cell' : missing_cells})

            missing_cells.to_csv('data_dash/missing_data.csv')
            risk_cell_dbm.to_csv('data_dash/risk_cell_dbm.csv')

            #df_data = pd.read_csv('data_dash/df_data.csv')
            print('OK data preparation')
            #df_loc = pd.read_csv('data_dash/df_loc.csv')
            df_loc.index = df_loc['Cell']
            df_loc = df_loc[['Longitude','Latitude']]

            def optimal_eps(data, neigh_points): #neighpoints is the number of neighbor points withing epsilon distance of point p + the point itself p
                neigh = NearestNeighbors(n_neighbors=neigh_points, algorithm='auto', p=2, metric='minkowski') #minkowski with p=2 is equivalent to euclidean distance
                nbrs = neigh.fit(data)
                distances, indices = nbrs.kneighbors(data)

                distances = np.sort(distances, axis=0)
                distances = distances[:, -1]

                i = 0 ; value = []
                for dist in distances :
                    i+=1
                    val = []
                    val.append(i)
                    val.append(dist)
                    value.append(val)

                rotor = Rotor()
                rotor.fit_rotate(value)
                elbow_index = rotor.get_elbow_index()
                for i in value :
                    if i[0] == elbow_index :
                        elbow_value = round(i[1],3)

                return elbow_index, elbow_value

            def PreClust(df_loc, neigh_points = [40,30,20]) :

                i = 0

                for neigh in neigh_points :
                    if i != -1 :
                        sc = StandardScaler()
                        df_loc_z = pd.DataFrame(sc.fit_transform(df_loc[['Latitude','Longitude']]))
                        df_loc_z.columns = ['Latitude','Longitude']

                        elbow_index, elbow_value = optimal_eps(df_loc, neigh)

                        db = DBSCAN(eps=elbow_value, min_samples=neigh, algorithm='auto', metric='euclidean').fit(df_loc[['Latitude','Longitude']])
                        labels = db.labels_
                        realClusterNum = len(set(labels)) - (1 if -1 in labels else 0)

                        if realClusterNum >= 3 :
                            df_loc["Cluster"] = [str(i+1) for i in labels]
                            i = -1
                        else :
                            if len(df_loc_z) > 10 :
                                k = 10
                            else :
                                k = len(df_loc_z)
                            kmeans = KMeans(n_clusters=k, n_init = 50, random_state = 3245).fit(df_loc_z[['Latitude','Longitude']])
                            pred = kmeans.predict(df_loc_z[['Latitude','Longitude']])
                            pred = pd.Series(pred, name='Cluster')
                            df_loc['Cluster'] = [str(i) for i in list(pred)]
                return df_loc

            df_loc = PreClust(df_loc)

            df_loc.to_csv('data_dash/df_localisation.csv')
            #df_loc = pd.read_csv('data_dash/df_localisation.csv') ; df_loc.index = df_loc.Cell ; df_loc['Cluster'] = [str(i) for i in df_loc.Cluster]
            fig = px.scatter(df_loc, x = 'Longitude', y = 'Latitude', color = 'Cluster')
            fig.update_layout(title_text="Localization pre-clusters")
            #print(fig['data']['markers']['color'])
            group = [] ; col = []
            for tuples in fig['data'] :
                group.append(tuples['legendgroup'])
                col.append(tuples['marker']['color'])

            styles = []
            for i in range(len(group)) :
                gr = group[i]
                co = col[i]
                line = '{loc_cluster} = '+str(gr)
                styles.append({
                'if': {
                    'filter_query': line,
                    'column_id': 'Cluster'
                },
                'color': '{}'.format(co)
            })

            #
            #df_data_1 = df_data.merge(df_loc['Cluster'], left_on = 'Cell', right_on = df_loc.index)
            #df_data_1.to_csv('data_dash/df_data.csv')
            #df_data = pd.read_csv('data_dash/df_data.csv')
            df_data = df_data.groupby('Cell').agg('mean')
            df_loc = df_data.merge(df_loc['Cluster'], left_on = 'Cell', right_on = df_loc.index)
            #df_loc['Cluster'] = [str(int(i)) for i in df_loc.Cluster]
            df_loc.to_csv('data_dash/df_geoloc.csv')

            df_data = pd.read_csv('data_dash/df_geoloc.csv')
            df_data['Cluster'] = [str(i) for i in df_data.Cluster]
            #df_data = df_data[df_data.Cluster == str(preclust['props']['value'])]

            sc = StandardScaler()
            df_loc_z = pd.DataFrame(sc.fit_transform(df_data[['RTWP','Avg Num HSUPA Users']]))
            df_loc_z.columns = ['RTWP','Avg Num HSUPA Users']

            kmeans_1 = KMeans(n_clusters=k_pc, random_state = 3245).fit(df_loc_z[['RTWP','Avg Num HSUPA Users']])

            pred_1 = kmeans_1.predict(df_loc_z[['RTWP','Avg Num HSUPA Users']])
            pred_1 = pd.Series(pred_1, name='Cluster')
            df_data['trafic_global'] = [str(i) for i in list(pred_1)]
            df_data['preclust'] = ['Cluster_'+str(df_data.Cluster[i])+str(df_data.trafic_global[i]) for i in df_data.index]
            df_preclusters = df_data
            df_preclusters.to_csv('data_dash/df_preclusters.csv')
            print('OK pre-clustering')
            df_loc = df_data
            df_data = pd.read_csv('data_dash/df_data.csv')
            df_data = df_data.merge(df_loc[['Cell','preclust','Cluster']], left_on = 'Cell', right_on = 'Cell')
            #df_data = pd.read_csv('data_dash/df_preclusters.csv')
            #table = df_data[['preclust','RTWP','Avg Num HSUPA Users']]
            table = pd.DataFrame()

            if detection_level == 'Soft' :
                rtwp = 0.95
                hsupa = 0.66
            elif detection_level == 'Average' :
                rtwp = 0.97
                hsupa = 0.5
            else :
                rtwp = 0.99
                hsupa = 0.33

            hsupa_quantile = df_data[['preclust','Avg Num HSUPA Users']].groupby('preclust').quantile(hsupa).sort_index() ;
            hsupa_quantile = round(hsupa_quantile, 2)
            # if rtwp == 'caps' :
            #     rtwp_quantile = []
            #     df_data = df_data.sort_values(by = 'preclust')
            #     for pc in df_data.preclust.unique() :
            #         rtwp_quantile.append(plt.boxplot(df_data[df_data.preclust == pc]['RTWP'])["caps"][1].get_data()[1][0].round(1))
            #     rtwp_quantile = pd.Series(rtwp_quantile)
            # else :
            rtwp_quantile = df_data[['preclust','RTWP']].groupby('preclust').quantile(rtwp).sort_index()
            rtwp_quantile = round(rtwp_quantile, 2)
            rtwp_quantile = list(rtwp_quantile['RTWP'])

            df_data = df_data.sort_values(by='preclust')
            preclusts = list(df_data.preclust.unique())
            cluster = [i[-2:-1] for i in preclusts]
            nb_values = df_loc['preclust'].value_counts().sort_index()

            table['Cluster'] = preclusts
            table['loc_cluster'] = cluster
            table['n_values'] = list(nb_values)
            table['RTWP threshold'] = rtwp_quantile
            table['Traffic threshold'] = list(hsupa_quantile['Avg Num HSUPA Users'])
            table = table[['Cluster','n_values','RTWP threshold','Traffic threshold','loc_cluster']]
            #table.index = table['loc_cluster']
            #table.to_csv('data_dash/table.csv')
            #table = round(table[['Traffic','n_values','RTWP','Avg Num HSUPA Users']],2) ; table['Traffic'] = [str(i) for i in table['Traffic']] ; table['n_values'] = [int(i) for i in table.n_values]
            output_table_pc = dash_table.DataTable(
                id='table',
                columns=[{"name": i, "id": i} for i in table.columns],
                data=table.to_dict('records'),
                #hidden_columns = ['loc_cluster'],
                style_header={
                    'color': 'black',
                    'fontWeight': 'bold',
                    'fontSize' : '15px'
                    },
                style_cell={'textAlign': 'center'},
                sort_action="native",
                style_table={
                'height': 450,
                'overflowY': 'auto'
                },
                style_data = {'width':'auto'},
                style_data_conditional=styles
                )

            #df_data = df_data.merge(df_preclust[['Cell','Cluster','trafic_global','preclust']], left_on = 'Cell', right_on = 'Cell')
            kept_columns = ['Cell','Time','RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)', '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic','preclust']
            colnames = ['Cell','Time','RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)', '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic','preclust']
            df = df_data[kept_columns] ; df.columns = colnames ; df.index = df.Cell
            # ACP et K-means
            KPI_COLUMNS_ACP = ['RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)', '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic']
            for i in df.preclust.unique() :
                  locals()['df_preclust_'+str(i)] = df[df.preclust == i]
                  # if rtwp == 'caps' :
                  #     locals()['RTWP_df_preclust_'+str(i)] = plt.boxplot(locals()['df_preclust_'+str(i)]['RTWP'])["caps"][1].get_data()[1][0].round(1)
                  # else :
                  locals()['RTWP_df_preclust_'+str(i)] = locals()['df_preclust_'+str(i)]['RTWP'].quantile(rtwp)
                  # if locals()['RTWP_df_preclust_'+str(i)] < -100 :
                  #     locals()['RTWP_df_preclust_'+str(i)] = -100
                  locals()['HSUPA_df_preclust_'+str(i)] = locals()['df_preclust_'+str(i)]['Avg Num HSUPA Users'].quantile(hsupa)
            sc = StandardScaler()
            acp = PCA(svd_solver='full')
            acp_cols = ['RTWP','Avg Num HSUPA Users','Avg Num HSDPA Users','Total Data Traffic Dl&UL(GBytes)','3G TRAFFIC SPEECH','3G&3G+ UL Traffic']
            for i in df.preclust.unique() :
                locals()['df_z_preclust_'+str(i)] = sc.fit_transform(locals()['df_preclust_'+str(i)][acp_cols])
                locals()['df_acp_preclust_'+str(i)] = pd.DataFrame(acp.fit_transform(locals()['df_z_preclust_'+str(i)]), index = locals()['df_preclust_'+str(i)].index, columns = ['Axe 1','Axe 2','Axe 3','Axe 4','Axe 5','Axe 6']).iloc[:,:4]
                if len(locals()['df_acp_preclust_'+str(i)]) < k_clust :
                    k_clust = len(locals()['df_acp_preclust_'+str(i)])
                kmeans = KMeans(n_clusters=k_clust, n_init = 30, random_state = 3245).fit(sc.fit_transform(locals()['df_acp_preclust_'+str(i)]))
                pred = kmeans.predict(sc.fit_transform(locals()['df_acp_preclust_'+str(i)]))
                pred = pd.Series(pred,name='Cluster')
                locals()['df_preclust_'+str(i)]['kmeans'] = [str(i) for i in list(pred)]
            for group in df.preclust.unique() :
                locals()['df_count_values_'+str(group)] = locals()['df_preclust_'+str(group)].kmeans.value_counts().sort_index()
                locals()['df_kmeans_'+str(group)] = locals()['df_preclust_'+str(group)].groupby('kmeans').agg('mean').sort_index()
                locals()['df_kmeans_'+str(group)]['n_obs'] = locals()['df_count_values_'+str(group)]
            for i in df.preclust.unique() :
                seuil_RTWP = int(locals()['RTWP_df_preclust_'+str(i)])
                seuil_HSUPA = float(locals()['HSUPA_df_preclust_'+str(i)])
                locals()['df_kmeans_'+str(i)]['High_RTWP'] = 0
                locals()['df_kmeans_'+str(i)]['Low_HSUPA'] = 0
                locals()['df_kmeans_'+str(i)]['Anom_risk'] = 0
                for k in locals()['df_kmeans_'+str(i)].index :
                    if locals()['df_kmeans_'+str(i)].RTWP[k] > seuil_RTWP :
                        locals()['df_kmeans_'+str(i)]['High_RTWP'][k] = 1
                    if locals()['df_kmeans_'+str(i)]['Avg Num HSUPA Users'][k] < seuil_HSUPA :
                        locals()['df_kmeans_'+str(i)]['Low_HSUPA'][k] = 1
                    if locals()['df_kmeans_'+str(i)]['Low_HSUPA'][k] == locals()['df_kmeans_'+str(i)]['High_RTWP'][k] == 1 :
                        locals()['df_kmeans_'+str(i)]['Anom_risk'][k] = 1
            for i in df.preclust.unique() :
                df_preclust = locals()['df_preclust_'+str(i)]
                df_preclust['kmeans'] = [str(i) for i in df_preclust.kmeans]
                df_kmeans = locals()['df_kmeans_'+str(i)]
                df_kmeans.index = [str(i) for i in df_kmeans.index]
                locals()['df_preclust_'+str(i)] = df_preclust.merge(df_kmeans['Anom_risk'], left_on = 'kmeans', right_on = df_kmeans.index)
                df_test = pd.DataFrame()
                for cell in locals()['df_preclust_'+str(i)].Cell.unique() :
                    df_cell = locals()['df_preclust_'+str(i)][locals()['df_preclust_'+str(i)].Cell == cell]
                    df_test = pd.concat([df_test, df_cell])
                    df_test.to_csv('data_dash/df_anom_risk_'+str(i)+'.csv')

            risk_prop = [] ; cell_risk = [] #; nb_cell = 0
            df_risk = pd.DataFrame()
            for i in df.preclust.unique() :
                locals()['df_preclust_'+str(i)] = pd.read_csv('data_dash/df_anom_risk_'+str(i)+'.csv')
                df_risk = pd.concat([df_risk, locals()['df_preclust_'+str(i)]])
                df_anom_risk = locals()['df_preclust_'+str(i)].groupby('Cell').agg('mean')
                for j in range(len(df_anom_risk)) :
                    risk_prop.append(df_anom_risk.Anom_risk[j])
                    cell_risk.append(df_anom_risk.index[j])
            risk_detection = pd.DataFrame({"Cell" : cell_risk, "prop" : risk_prop}).round(3)
            risk_cell = [1 if i > riskobs/100 else 0 for i in risk_detection.prop]
            risk_detection['risk_cell'] = risk_cell
            #nb_risk = len(risk_detection[risk_detection.risk_cell == 1])
            #nb_cell = len(risk_detection)
            #prop_risk = round(nb_risk*100/nb_cell,2)

            df_risk.to_csv('data_dash/df_risk.csv')
            risk_detection.to_csv('data_dash/risk_detection.csv')

            df = pd.read_csv('data_dash/df_clean.csv')
            df['Time'] = pd.to_datetime(df['Time'])
            #risk_detection = pd.read_csv('risk_detection.csv')
            #risk_cell_dbm = pd.read_csv('data_dash/risk_cell_dbm.csv')
            print('OK clustering')

            #risk_detection = pd.read_csv('risk_detection.csv')
            #del risk_detection['anomaly']
            #del risk_detection['final_ranking']
            #df_risk = pd.read_csv('df_risk.csv')

            risk_cells = list(risk_detection[risk_detection.risk_cell == 1].Cell)
            for cell in list(risk_cell_dbm.Cell) :
                risk_cells.append(cell)

            data_adtk = pd.DataFrame()
            for cell in list(risk_cells) :
                df_pc = df[df.Cell == cell]
                df_pc = df_pc.sort_values(by = 'Time')
                df_pc['Time'] = pd.to_datetime(df_pc['Time'])
                time = min(df_pc.Time)
                new_time = []
                for i in range(1,len(df_pc)+1) :
                    new_time.append(time+timedelta(hours=i))
                df_pc['new_time'] = new_time
                data_adtk = pd.concat([data_adtk,df_pc])

            data_adtk['real_time'] = data_adtk['Time']
            data_adtk['Time'] = data_adtk['new_time']

            anomaly_detection = pd.DataFrame(data = {'Cell' : risk_cells})

            nb_cells = len(anomaly_detection)

            def InterfDetec(persist_ad = True, volatilityshift_ad = True, levelshift_ad = True, c_persist = 2, window_persist = 1, c_levelshift = 2, window_levelshift = 15, c_volatility = 9, window_volatility = 20, side = 'both'):
                df.loc[:,'Time']=pd.to_datetime(df.loc[:,'Time'], format="%Y-%m-%d %H:%M:%S")
                df.index = df.Time
                if persist_ad :
                    persist_ad = PersistAD(c=c_persist, side=side,window=window_persist)
                    for cell in list(anomaly_detection.Cell) :
                        df_anom = df[df.Cell == cell] ;
                        if len(df_anom) > (2*window_persist + 10) :
                            anomalies = persist_ad.fit_detect(df_anom.RTWP)
                            persist_hour[cell] = anomalies[anomalies == 1]
                            persist_anom[cell] = 1 if len(anomalies[anomalies == 1]) > 0 else 0
                if levelshift_ad :
                    ls_ad = PersistAD(c=c_levelshift, side=side,window=window_levelshift)
                    for cell in list(anomaly_detection.Cell) :
                        df_anom = df[df.Cell == cell] ;
                        if len(df_anom) > (2*window_levelshift + 10) :
                            anomalies = ls_ad.fit_detect(df_anom.RTWP)
                            levelshift_hour[cell] = anomalies[anomalies == 1]
                            levelshift_anom[cell] = 1 if len(anomalies[anomalies == 1]) > 0 else 0
                if volatilityshift_ad :
                    vs_ad = VolatilityShiftAD(c=c_volatility, side=side,window=window_volatility)
                    for cell in list(anomaly_detection.Cell) :
                        df_anom = df[df.Cell == cell] ;
                        if len(df_anom) > (2*window_volatility + 10) :
                            anomalies = vs_ad.fit_detect(df_anom.RTWP)
                            volatility_hour[cell] = anomalies[anomalies == 1]
                            volatility_anom[cell] = 1 if len(anomalies[anomalies == 1]) > 0 else 0

            df = data_adtk
            persist_anom = {} ; levelshift_anom = {} ; volatility_anom = {}
            persist_hour = {} ; levelshift_hour = {} ; volatility_hour = {}

            InterfDetec(c_persist = 7, window_persist = 1, c_levelshift = 4, window_levelshift = 6, c_volatility = 5, window_volatility = 24)

            persist_anom = pd.DataFrame(pd.Series(persist_anom), columns = ['basic_outlier'])
            levelshift_anom = pd.DataFrame(pd.Series(levelshift_anom), columns = ['basic_levelshift'])
            volatilityshift_anom = pd.DataFrame(pd.Series(volatility_anom), columns = ['basic_volatilityshift'])

            if 'basic_outlier' in anomaly_detection.columns :
                del anomaly_detection['basic_outlier']
            if 'basic_levelshift' in anomaly_detection.columns :
                del anomaly_detection['basic_levelshift']
            if 'basic_volatilityshift' in anomaly_detection.columns :
                del anomaly_detection['basic_volatilityshift']

            anomaly_detection = anomaly_detection.merge(persist_anom, left_on = 'Cell', right_on = persist_anom.index)
            anomaly_detection = anomaly_detection.merge(levelshift_anom, left_on = 'Cell', right_on = levelshift_anom.index)
            anomaly_detection = anomaly_detection.merge(volatilityshift_anom, left_on = 'Cell', right_on = volatilityshift_anom.index)
            anomaly_detection['basic_anomaly'] = [1 if anomaly_detection.basic_outlier[i]+anomaly_detection.basic_levelshift[i]+anomaly_detection.basic_volatilityshift[i] > 0 else 0 for i in range(len(anomaly_detection))]
            anomaly_detection['basic_all_anomaly'] = [1 if anomaly_detection.basic_outlier[i]+anomaly_detection.basic_levelshift[i]+anomaly_detection.basic_volatilityshift[i] == 3 else 0 for i in range(len(anomaly_detection))]

            df = data_adtk
            persist_anom = {} ; levelshift_anom = {} ; volatility_anom = {}
            persist_hour = {} ; levelshift_hour = {} ; volatility_hour = {}

            # InterfDetec(c_persist = 21, window_persist = 1, c_levelshift = 12, window_levelshift = 6, c_volatility = 15, window_volatility = 24)
            #
            # persist_anom = pd.DataFrame(pd.Series(persist_anom), columns = ['high_outlier'])
            # levelshift_anom = pd.DataFrame(pd.Series(levelshift_anom), columns = ['high_levelshift'])
            # volatilityshift_anom = pd.DataFrame(pd.Series(volatility_anom), columns = ['high_volatilityshift'])
            #
            # if 'high_outlier' in anomaly_detection.columns :
            #     del anomaly_detection['high_outlier']
            # if 'high_levelshift' in anomaly_detection.columns :
            #     del anomaly_detection['high_levelshift']
            # if 'high_volatilityshift' in anomaly_detection.columns :
            #     del anomaly_detection['high_volatilityshift']
            #
            #
            # anomaly_detection = anomaly_detection.merge(persist_anom, left_on = 'Cell', right_on = persist_anom.index)
            # anomaly_detection = anomaly_detection.merge(levelshift_anom, left_on = 'Cell', right_on = levelshift_anom.index)
            # anomaly_detection = anomaly_detection.merge(volatilityshift_anom, left_on = 'Cell', right_on = volatilityshift_anom.index)
            # anomaly_detection['high_anomaly'] = [1 if anomaly_detection.high_outlier[i]+anomaly_detection.high_levelshift[i]+anomaly_detection.high_volatilityshift[i] > 0 else 0 for i in range(len(anomaly_detection))]
            # anomaly_detection['high_all_anomaly'] = [1 if anomaly_detection.high_outlier[i]+anomaly_detection.high_levelshift[i]+anomaly_detection.high_volatilityshift[i] == 3 else 0 for i in range(len(anomaly_detection))]

            anomaly_detection.to_csv('data_dash/anomaly_detection.csv')
            nb_anom = len(anomaly_detection[anomaly_detection.basic_anomaly == 1])
            anom_cell = []
            for cell in risk_detection.Cell :
                if cell in list(anomaly_detection[anomaly_detection.basic_anomaly == 1].Cell) :
                    anom_cell.append(1)
                else :
                    anom_cell.append(0)
            risk_detection['anomaly'] = anom_cell

            risk_detection.to_csv('data_dash/risk_detection.csv')
            data_adtk.to_csv('data_dash/data_adtk.csv')
            #prop_anom = round(nb_anom*100/nb_cells,2)
            print('OK anomaly detection')

            # anomaly_detection = pd.read_csv('data_dash/anomaly_detection.csv')
            # data_adtk = pd.read_csv('data_dash/data_adtk.csv')
            # df_risk = pd.read_csv('data_dash/df_risk.csv')
            # new_df = pd.read_csv('data_dash/new_df.csv')
            # df_preclusters = pd.read_csv('data_dash/df_preclusters.csv')
            # df_clean = pd.read_csv('data_dash/df_clean.csv')
            # df_clean['DCR-PS'] = df_clean['3G DCR ALL DATA SERVICES']


            #data_adtk = pd.read_csv('data_dash/data_adtk.csv')
            data_adtk['Time'] = pd.to_datetime(data_adtk['Time'])
            #anomaly_detection = pd.read_csv('data_dash/anomaly_detection.csv')
            data_adtk['real_time_str'] = [str(i) for i in data_adtk['real_time']]
            df_risk['Time_str'] = [str(i) for i in df_risk['Time']]
            data_adtk = data_adtk.merge(df_risk[['Cell', 'Time_str', 'Anom_risk']], how='left',
                                        left_on=['Cell', 'real_time_str'], right_on=['Cell', 'Time_str'])

            anomaly_detection = anomaly_detection[anomaly_detection.basic_anomaly == 1]
            #anom_detec = list(anomaly_detection.Cell)
            #data_ad = len(data_adtk)
            def get_delta(d1, d2):
                delta = d2 - d1
                return delta

            anom_cells = []
            basic_last_anom = []
            basic_anom_last = []
            # high_last_anom = []
            # high_anom_last = []
            # new_high_anom = []
            maximum_RTWP = []
            new_df = pd.DataFrame()

            len_new_df = [] ; len_new_df_dropna = []

            for cell in anomaly_detection.Cell:
                df_cell = data_adtk[data_adtk.Cell == cell]
                df_cell.index = pd.to_datetime(df_cell.Time)
                df_cell = df_cell.sort_index()

                # basic anomalies
                all_anomalies = []
                persistad = PersistAD(c = 7, window = 1, side = 'both')
                anomalies = persistad.fit_detect(df_cell['RTWP'])
                all_anomalies.append(anomalies[anomalies == 1].index)
                persistad = PersistAD(c = 4, window = 6, side = 'both')
                anomalies = persistad.fit_detect(df_cell['RTWP'])
                all_anomalies.append(anomalies[anomalies == 1].index)
                persistad = VolatilityShiftAD(c = 5, window = 24, side = 'both')
                anomalies = persistad.fit_detect(df_cell['RTWP'])
                all_anomalies.append(anomalies[anomalies == 1].index)

                anomalies = []
                for hours in all_anomalies:
                    for hour in hours:
                        anomalies.append(hour)
                anomalies = list(set(anomalies))

                all_anomalies = sorted(anomalies)

                anom_range_final = []
                i = 0
                while i != len(all_anomalies):
                    anom_range = []
                    flag = 1
                    while flag == 1:
                        anom = all_anomalies[i]
                        if i < len(all_anomalies) - 1:
                            if all_anomalies[i + 1] < anom + timedelta(hours=24):
                                anom_range.append(anom)
                                i += 1
                            else:
                                anom_range.append(anom)
                                flag = 0
                                i += 1
                                anom_range_final.append(anom_range)
                        else:
                            anom_range.append(anom)
                            flag = 0
                            i += 1
                            anom_range_final.append(anom_range)

                new_time_range = []
                for time_range in anom_range_final:
                    if len(time_range) > 1:
                        new_time_range.append([time_range[0], time_range[-1]])
                    else:
                        newnew = [time_range[0], time_range[0]]
                        new_time_range.append(newnew)

                mean_rtwp = df_cell['RTWP'].mean()
                std_rtwp = df_cell['RTWP'].std()
                newnew_time_range = []
                for time_range in new_time_range:
                    flag_low = 1
                    flag_high = 1
                    if len(time_range) > 1:
                        while flag_low == 1:
                            low_time = time_range[0] - timedelta(hours=12)
                            df_before = df_cell[(df_cell.Time < time_range[0]) & (df_cell.Time > low_time)]
                            rtwp_before = df_before['RTWP'].mean()
                            std_before = df_before['RTWP'].std()
                            if (rtwp_before > mean_rtwp + 3) or (std_before > 2 * std_rtwp):
                                time_range[0] = low_time
                            else:
                                flag_low = 0

                        while flag_high == 1:
                            high_time = time_range[1] + timedelta(hours=12)
                            df_after = df_cell[(df_cell.Time > time_range[1]) & (df_cell.Time < high_time)]
                            rtwp_after = df_after['RTWP'].mean()
                            std_after = df_after['RTWP'].std()
                            if (rtwp_after > mean_rtwp + 3) or (std_after > 2 * std_rtwp):
                                time_range[1] = high_time
                            else:
                                flag_high = 0
                    newnew_time_range.append(time_range)

                df_cell['basic_anom'] = 0
                for timerange in newnew_time_range:
                    df_range = df_cell[(df_cell.Time >= timerange[0] + timedelta(hours=-3)) & (
                                df_cell.Time <= timerange[1] + timedelta(hours=3))]
                    if sum(df_range['Anom_risk']) > 0:
                        delta = get_delta(timerange[0], timerange[1])
                        for i in range(delta.days * 24 + delta.seconds // 3600 + 1):
                            hour = timerange[0] + timedelta(hours=i)
                            df_cell.loc[hour, 'basic_anom'] = 1

                df_normal = df_cell[df_cell.basic_anom == 0]
                mean_rtwp = df_normal['RTWP'].mean()
                std_rtwp = df_normal['RTWP'].std()
                newnewnew_time_range = []
                for time_range in newnew_time_range:
                    flag_low = 1
                    flag_high = 1
                    if len(time_range) > 1:
                        while flag_low == 1:
                            low_time = time_range[0] - timedelta(hours=12)
                            df_before = df_cell[(df_cell.Time < time_range[0]) & (df_cell.Time > low_time)]
                            rtwp_before = df_before['RTWP'].mean()
                            std_before = df_before['RTWP'].std()
                            if (rtwp_before > mean_rtwp + 3) or (std_before > 2 * std_rtwp):
                                time_range[0] = low_time
                            else:
                                flag_low = 0

                        while flag_high == 1:
                            high_time = time_range[1] + timedelta(hours=12)
                            df_after = df_cell[(df_cell.Time > time_range[1]) & (df_cell.Time < high_time)]
                            rtwp_after = df_after['RTWP'].mean()
                            std_after = df_after['RTWP'].std()
                            if (rtwp_after > mean_rtwp + 3) or (std_after > 2 * std_rtwp):
                                time_range[1] = high_time
                            else:
                                flag_high = 0
                    newnewnew_time_range.append(time_range)

                risk_timerange = []
                for timerange in newnewnew_time_range:
                    df_range = df_cell[(df_cell.Time >= timerange[0] + timedelta(hours=-3)) & (
                                df_cell.Time <= timerange[1] + timedelta(hours=3))]
                    if sum(df_range['Anom_risk']) > 0:
                        delta = get_delta(timerange[0], timerange[1])
                        for i in range(delta.days * 24 + delta.seconds // 3600 + 1):
                            hour = timerange[0] + timedelta(hours=i)
                            df_cell.loc[hour, 'basic_anom'] = 1
                        risk_timerange.append(timerange)
                newnewnew_time_range = risk_timerange

                newnewnew_time_range = sorted(newnewnew_time_range, key=itemgetter(-1))
                if len(newnewnew_time_range) > 0:
                    last_anomaly = newnewnew_time_range[-1]
                    if last_anomaly[1] > max(df_cell.Time):
                        last_anomaly[1] = max(df_cell.Time)
                    delta = get_delta(last_anomaly[0], last_anomaly[1])
                    anom_cells.append(cell);
                    basic_last_anom.append(last_anomaly[-1]);
                    basic_anom_last.append(delta)
                else:
                    anom_cells.append('NaN')
                    basic_last_anom.append('NaN');
                    basic_anom_last.append('NaN')

                df_cell_basic = df_cell

                new_df = pd.concat([new_df, df_cell])

                if len(df_cell[df_cell.basic_anom == 1]) > 0 :
                    maximum_RTWP.append(max(df_cell[df_cell.basic_anom == 1].RTWP))
                else :
                    maximum_RTWP.append(0)

            new_df = new_df.dropna()

            anom_kpi = pd.DataFrame(
                data={'Cell': anom_cells, 'Since_anom_end': basic_last_anom, 'Anom_duration': basic_anom_last, 'max_RTWP': maximum_RTWP})
            anom_kpi = anom_kpi[anom_kpi.Cell != 'NaN']

            real_last_basic_anom = []
            # for i in range(len(anom_kpi)):
            for i in anom_kpi.index:
                last_anom = anom_kpi.loc[i, 'Since_anom_end']
                df_cell = data_adtk[data_adtk.Cell == anom_kpi.loc[i, 'Cell']]
                if last_anom in list(df_cell.Time):
                    real = max(df_cell[df_cell.Time == last_anom]['real_time'])
                else:
                    real = last_anom
                real_last_basic_anom.append(real)
            anom_kpi['Since_anom_end_real'] = real_last_basic_anom

            last_basic_anomaly = []
            last_high_anomaly = []
            #
            # for i in range(len(anom_kpi)):
            for i in anom_kpi.index:
                df_cell = data_adtk[data_adtk.Cell == anom_kpi.loc[i, 'Cell']];
                df_cell.real_time = pd.to_datetime(df_cell.real_time)
                max_time = max(df_cell.real_time)
                basic = pd.to_datetime(anom_kpi.loc[i, 'Since_anom_end_real'])
                last_basic_anomaly.append(get_delta(basic, max_time))

            anom_kpi['Since_anom_end'] = last_basic_anomaly

            if 'DCR-PS' not in new_df.columns and '3G DCR ALL DATA SERVICES' in new_df.columns:
                new_df['DCR-PS'] = new_df['3G DCR ALL DATA SERVICES']


            cells = [] ; basic_dcr = [] ; basic_cssrcs = [] ; basic_cssrps = [] ; high_dcr = [] ; high_cssrcs = [] ; high_cssrps = []

            for cell in new_df.Cell.unique():
                if 'DCR-PS' in new_df.columns or 'CSSR-CS' in new_df.columns or 'CSSR-PS' in new_df.columns:
                    cells.append(cell)
                df_cell_basic = new_df[(new_df.Cell == cell) & (new_df.basic_anom == 1)].dropna()
                df_cell_normal = new_df[(new_df.Cell == cell) & (new_df.basic_anom == 0)].dropna()
                if 'DCR-PS' in new_df.columns:
                    basic_dcr.append(
                        round(df_cell_basic['DCR-PS'].mean() - df_cell_normal['DCR-PS'].mean(), 2) * poids_DCRPS)
                if 'CSSR-CS' in new_df.columns:
                    basic_cssrcs.append(
                        round(df_cell_normal['CSSR-CS'].mean() - df_cell_basic['CSSR-CS'].mean(), 2) * poids_CSSRCS)
                if 'CSSR-PS' in new_df.columns:
                    basic_cssrps.append(
                        round(df_cell_normal['CSSR-PS'].mean() - df_cell_basic['CSSR-PS'].mean(), 2) * poids_CSSRPS)

            if ('DCR-PS' in new_df.columns and 'DCR-PS' in qos_kpi) and (
                    'CSSR-CS' in new_df.columns and 'CSSR-CS' in qos_kpi) and (
                    'CSSR-PS' in new_df.columns and 'CSSR-PS' in qos_kpi):
                df_anom = pd.DataFrame(data={'Cell': cells, 'basic_DCR_PS': basic_dcr, 'basic_CSSR_CS': basic_cssrcs,
                                             'basic_CSSR_PS': basic_cssrps})
                poids = poids_DCRPS + poids_CSSRCS + poids_CSSRPS
                df_anom['basic_sum_degr'] = [sum(df_anom.iloc[i, [1, 2, 3]]) / poids for i in range(len(df_anom))]

            elif ('DCR-PS' in new_df.columns and 'DCR-PS' in qos_kpi) and (
                    'CSSR-CS' in new_df.columns and 'CSSR-CS' in qos_kpi) and (
                    'CSSR-PS' not in new_df.columns or 'CSSR-PS' not in qos_kpi):
                df_anom = pd.DataFrame(data={'Cell': cells, 'basic_DCR_PS': basic_dcr, 'basic_CSSR_CS': basic_cssrcs})
                poids = poids_DCRPS + poids_CSSRCS
                df_anom['basic_sum_degr'] = [sum(df_anom.iloc[i, [1, 2]]) / poids for i in range(len(df_anom))]

            elif ('DCR-PS' in new_df.columns and 'DCR-PS' in qos_kpi) and (
                    'CSSR-CS' not in new_df.columns or 'CSSR-CS' not in qos_kpi) and (
                    'CSSR-PS' in new_df.columns and 'CSSR-PS' in qos_kpi):
                df_anom = pd.DataFrame(data={'Cell': cells, 'basic_DCR_PS': basic_dcr, 'basic_CSSR_PS': basic_cssrps})
                poids = poids_DCRPS + poids_CSSRPS
                df_anom['basic_sum_degr'] = [sum(df_anom.iloc[i, [1, 2]]) / poids for i in range(len(df_anom))]

            elif ('DCR-PS' not in new_df.columns or 'DCR-PS' not in qos_kpi) and (
                    'CSSR-CS' in new_df.columns and 'CSSR-CS' in qos_kpi) and (
                    'CSSR-PS' in new_df.columns and 'CSSR-PS' in qos_kpi):
                df_anom = pd.DataFrame(data={'Cell': cells, 'basic_CSSR_CS': basic_cssrcs, 'basic_CSSR_PS': basic_cssrps})
                poids = poids_CSSRCS + poids_CSSRPS
                df_anom['basic_sum_degr'] = [sum(df_anom.iloc[i, [1, 2]]) / poids for i in range(len(df_anom))]

            elif ('DCR-PS' not in new_df.columns or 'DCR-PS' not in qos_kpi) and (
                    'CSSR-CS' not in new_df.columns or 'CSSR-CS' not in qos_kpi) and (
                    'CSSR-PS' in new_df.columns and 'CSSR-PS' in qos_kpi):
                df_anom = pd.DataFrame(data={'Cell': cells, 'basic_CSSR_PS': basic_cssrps})
                df_anom['basic_sum_degr'] = df_anom['basic_CSSR_PS'] / poids_CSSRPS

            elif ('DCR-PS' not in new_df.columns or 'DCR-PS' not in qos_kpi) and (
                    'CSSR-CS' in new_df.columns and 'CSSR-CS' in qos_kpi) and (
                    'CSSR-PS' not in new_df.columns or 'CSSR-PS' not in qos_kpi):
                df_anom = pd.DataFrame(data={'Cell': cells, 'basic_CSSR_CS': basic_cssrcs})
                df_anom['basic_sum_degr'] = df_anom['basic_CSSR_CS'] / poids_CSSRCS

            elif ('DCR-PS' in new_df.columns and 'DCR-PS' in qos_kpi) and (
                    'CSSR-CS' not in new_df.columns or 'CSSR-CS' not in qos_kpi) and (
                    'CSSR-PS' not in new_df.columns or 'CSSR-PS' not in qos_kpi):
                df_anom = pd.DataFrame(data={'Cell': cells, 'basic_DCR_PS': basic_dcr})
                df_anom['basic_sum_degr'] = df_anom['basic_DCR_PS'] / poids_DCRPS

            else :
                df_anom = pd.DataFrame(data = {'Cell' : cells})
                df_anom['basic_sum_degr'] = 0

            df_anom = df_anom.merge(anom_kpi, how='inner', left_on = 'Cell', right_on = 'Cell')

            anom = anomaly_detection.merge(df_anom, how = 'inner', left_on = 'Cell', right_on = 'Cell')
            anom['qos_degradation'] = anom['basic_sum_degr']

            anom['ranking'] = ''

            for i in anom.index :
                if 'max RTWP' in indicators :
                    if anom.loc[i, 'max_RTWP'] > -85:
                        anom['ranking'][i] += '1'
                    elif anom.loc[i, 'max_RTWP'] > -90:
                        anom['ranking'][i] += '2'
                    if anom.loc[i, 'max_RTWP'] > -95:
                        anom['ranking'][i] += '3'
                    if anom.loc[i, 'max_RTWP'] > -100:
                        anom['ranking'][i] += '4'
                    else:
                        anom['ranking'][i] += '5'

                if 'last anomaly' in indicators:
                    if anom.loc[i, 'Since_anom_end'] < timedelta(days=3):
                        anom['ranking'][i] += '1'
                    elif anom.loc[i, 'Since_anom_end'] < timedelta(days=7):
                        anom['ranking'][i] += '2'
                    else:
                        anom['ranking'][i] += '3'

                if 'anomaly duration' in indicators:
                    if int(str(anom.loc[i, 'Anom_duration']).split(' ')[0]) > 2:
                        anom['ranking'][i] += '1'
                    else:
                        anom['ranking'][i] += '2'

            if 'qos degradation' in indicators:
                anom = anom.sort_values(by=['ranking', 'qos_degradation'], ascending=[True, False])
            else:
                anom = anom.sort_values(by='ranking', ascending=True)

            i = 0
            anom['final_ranking'] = 0
            for line in anom.index :
                i+=1
                anom['final_ranking'][line] = i

            risk_detection = pd.read_csv('data_dash/risk_detection.csv')
            if 'final_ranking' in list(risk_detection.columns) :
                del risk_detection['final_ranking']

            risk_detection = risk_detection.merge(anom[['Cell','final_ranking']], how = 'left', left_on = 'Cell', right_on = 'Cell')
            risk_detection = risk_detection.sort_values(by = 'final_ranking')

            risk_detection.to_csv('data_dash/risk_detection.csv')

            anom.to_csv('data_dash/anom.csv')
            new_df.to_csv('data_dash/new_df.csv')
            print('OK anomaly prioritization')

            main_output = pd.DataFrame()
            anom['last_anomaly'] = anom['Since_anom_end']
            anom['last_anomaly_range'] = anom['Anom_duration']
            main_output = anom[['Cell','last_anomaly','last_anomaly_range','final_ranking']].sort_values(by = 'final_ranking')
            #main_output.index = main_output.final_ranking
            #main_output = main_output.sort_index()
            main_output = main_output.merge(df_preclusters[['Cell','preclust']], left_on = 'Cell', right_on = 'Cell')
            main_output = main_output[['final_ranking','Cell','preclust','last_anomaly','last_anomaly_range']]

            for line in main_output.index :
                if main_output.loc[line,'last_anomaly_range'] == '0 days 00:00:00' :
                    main_output.loc[line,'last_anomaly_range'] = '0 days 01:00:00'

            main_output_vide = pd.DataFrame()

            main_output_table = dash_table.DataTable(
                id='main_output_table',
                columns=[{"name": i, "id": i} for i in main_output_vide.columns],
                data=main_output_vide.to_dict('records'),
                style_header={
                    'color': 'black',
                    'fontWeight': 'bold',
                    'fontSize' : '15px'
                    },
                style_cell={'textAlign': 'center'},
                sort_action="native",
                style_table={
                'height': 450,
                'overflowY': 'auto'
                },
                style_data = {'width':'auto'}
                )

            main_output_table = dash_table.DataTable(
                id='main_output_table',
                columns=[{"name": i, "id": i} for i in main_output.columns],
                data=main_output.to_dict('records'),
                style_header={
                    'color': 'black',
                    'fontWeight': 'bold',
                    'fontSize' : '15px'
                    },
                style_cell={'textAlign': 'center'},
                sort_action="native",
                style_table={
                'height': 450,
                'overflowY': 'auto'
                },
                style_data = {'width':'auto'}
                )

            main_output.to_csv('data_dash/main_output.csv')

            #df_main_output = main_output.values.tolist()

            if 'DCR-PS' in new_df.columns and 'CSSR-CS' not in new_df.columns and 'CSSR-PS' not in new_df.columns :
                new_df = new_df[['Cell', 'RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)',
                             '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic', 'DCR-PS', 'basic_anom']]
                columns = ['RTWP_non_anom','RTWP_anom', 'Avg Num HSUPA Users_non_anom', 'Avg Num HSUPA Users_anom', 'Avg Num HSDPA Users_non_anom', 'Avg Num HSDPA Users_anom', 'Total Data Traffic Dl&UL(GBytes)_non_anom', 'Total Data Traffic Dl&UL(GBytes)_anom', '3G TRAFFIC SPEECH_non_anom', '3G TRAFFIC SPEECH_anom',
                           '3G&3G+ UL Traffic_non_anom', '3G&3G+ UL Traffic_anom' , 'DCR-PS_non_anom', 'DCR-PS_anom']

            if 'DCR-PS' not in new_df.columns and 'CSSR-CS' in new_df.columns and 'CSSR-PS' not in new_df.columns :
                new_df = new_df[['Cell', 'RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)',
                             '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic', 'CSSR-CS', 'basic_anom']]
                columns = ['RTWP_non_anom','RTWP_anom', 'Avg Num HSUPA Users_non_anom', 'Avg Num HSUPA Users_anom', 'Avg Num HSDPA Users_non_anom', 'Avg Num HSDPA Users_anom', 'Total Data Traffic Dl&UL(GBytes)_non_anom', 'Total Data Traffic Dl&UL(GBytes)_anom', '3G TRAFFIC SPEECH_non_anom', '3G TRAFFIC SPEECH_anom',
                           '3G&3G+ UL Traffic_non_anom', '3G&3G+ UL Traffic_anom' , 'CSSR-CS_non_anom', 'CSSR-CS_anom']

            if 'DCR-PS' not in new_df.columns and 'CSSR-CS' not in new_df.columns and 'CSSR-PS' in new_df.columns :
                new_df = new_df[['Cell', 'RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)',
                             '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic', 'CSSR-PS', 'basic_anom']]
                columns = ['RTWP_non_anom','RTWP_anom', 'Avg Num HSUPA Users_non_anom', 'Avg Num HSUPA Users_anom', 'Avg Num HSDPA Users_non_anom', 'Avg Num HSDPA Users_anom', 'Total Data Traffic Dl&UL(GBytes)_non_anom', 'Total Data Traffic Dl&UL(GBytes)_anom', '3G TRAFFIC SPEECH_non_anom', '3G TRAFFIC SPEECH_anom',
                           '3G&3G+ UL Traffic_non_anom', '3G&3G+ UL Traffic_anom' , 'CSSR-PS_non_anom', 'CSSR-PS_anom']

            if 'DCR-PS' in new_df.columns and 'CSSR-CS' in new_df.columns and 'CSSR-PS' not in new_df.columns :
                new_df = new_df[['Cell', 'RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)',
                             '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic', 'DCR-PS', 'CSSR-CS', 'basic_anom']]
                columns = ['RTWP_non_anom','RTWP_anom', 'Avg Num HSUPA Users_non_anom', 'Avg Num HSUPA Users_anom', 'Avg Num HSDPA Users_non_anom', 'Avg Num HSDPA Users_anom', 'Total Data Traffic Dl&UL(GBytes)_non_anom', 'Total Data Traffic Dl&UL(GBytes)_anom', '3G TRAFFIC SPEECH_non_anom', '3G TRAFFIC SPEECH_anom',
                           '3G&3G+ UL Traffic_non_anom', '3G&3G+ UL Traffic_anom' , 'DCR-PS_non_anom', 'DCR-PS_anom', 'CSSR-CS_non_anom', 'CSSR-CS_anom']

            if 'DCR-PS' in new_df.columns and 'CSSR-CS' not in new_df.columns and 'CSSR-PS' in new_df.columns :
                new_df = new_df[['Cell', 'RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)',
                             '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic', 'DCR-PS', 'CSSR-PS', 'basic_anom']]
                columns = ['RTWP_non_anom','RTWP_anom', 'Avg Num HSUPA Users_non_anom', 'Avg Num HSUPA Users_anom', 'Avg Num HSDPA Users_non_anom', 'Avg Num HSDPA Users_anom', 'Total Data Traffic Dl&UL(GBytes)_non_anom', 'Total Data Traffic Dl&UL(GBytes)_anom', '3G TRAFFIC SPEECH_non_anom', '3G TRAFFIC SPEECH_anom',
                           '3G&3G+ UL Traffic_non_anom', '3G&3G+ UL Traffic_anom' , 'DCR-PS_non_anom', 'DCR-PS_anom', 'CSSR-PS_non_anom', 'CSSR-PS_anom']

            if 'DCR-PS' not in new_df.columns and 'CSSR-CS' in new_df.columns and 'CSSR-PS' in new_df.columns :
                new_df = new_df[['Cell', 'RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)',
                             '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic', 'CSSR-PS', 'CSSR-CS', 'basic_anom']]
                columns = ['RTWP_non_anom','RTWP_anom', 'Avg Num HSUPA Users_non_anom', 'Avg Num HSUPA Users_anom', 'Avg Num HSDPA Users_non_anom', 'Avg Num HSDPA Users_anom', 'Total Data Traffic Dl&UL(GBytes)_non_anom', 'Total Data Traffic Dl&UL(GBytes)_anom', '3G TRAFFIC SPEECH_non_anom', '3G TRAFFIC SPEECH_anom',
                           '3G&3G+ UL Traffic_non_anom', '3G&3G+ UL Traffic_anom' , 'CSSR-PS_non_anom', 'CSSR-PS_anom', 'CSSR-CS_non_anom', 'CSSR-CS_anom']

            if 'DCR-PS' in new_df.columns and 'CSSR-CS' in new_df.columns and 'CSSR-PS' in new_df.columns :
                new_df = new_df[['Cell', 'RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)',
                             '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic', 'DCR-PS', 'CSSR-CS', 'CSSR-PS', 'basic_anom']]
                columns = ['RTWP_non_anom','RTWP_anom', 'Avg Num HSUPA Users_non_anom', 'Avg Num HSUPA Users_anom', 'Avg Num HSDPA Users_non_anom', 'Avg Num HSDPA Users_anom', 'Total Data Traffic Dl&UL(GBytes)_non_anom', 'Total Data Traffic Dl&UL(GBytes)_anom', '3G TRAFFIC SPEECH_non_anom', '3G TRAFFIC SPEECH_anom',
                           '3G&3G+ UL Traffic_non_anom', '3G&3G+ UL Traffic_anom' , 'DCR-PS_non_anom', 'DCR-PS_anom', 'CSSR-CS_non_anom', 'CSSR-CS_anom', 'CSSR-PS_non_anom', 'CSSR-PS_anom']

            if 'DCR-PS' not in new_df.columns and 'CSSR-CS' not in new_df.columns and 'CSSR-PS' not in new_df.columns :
                new_df = new_df[['Cell', 'RTWP', 'Avg Num HSUPA Users', 'Avg Num HSDPA Users', 'Total Data Traffic Dl&UL(GBytes)',
                             '3G TRAFFIC SPEECH', '3G&3G+ UL Traffic', 'basic_anom']]
                columns = ['RTWP_non_anom','RTWP_anom', 'Avg Num HSUPA Users_non_anom', 'Avg Num HSUPA Users_anom', 'Avg Num HSDPA Users_non_anom', 'Avg Num HSDPA Users_anom', 'Total Data Traffic Dl&UL(GBytes)_non_anom', 'Total Data Traffic Dl&UL(GBytes)_anom', '3G TRAFFIC SPEECH_non_anom', '3G TRAFFIC SPEECH_anom',
                           '3G&3G+ UL Traffic_non_anom', '3G&3G+ UL Traffic_anom']



            #detailled_output
            basic_anom_obs = new_df[new_df.basic_anom == 1] ; non_basic_anom_obs = new_df[new_df.basic_anom == 0]
            group_basic_anom_obs = basic_anom_obs.groupby('Cell').agg('mean') ; group_basic_anom_obs.columns = [i+'_anom' for i in group_basic_anom_obs.columns]
            group_non_basic_anom_obs = non_basic_anom_obs.groupby('Cell').agg('mean') ; group_non_basic_anom_obs.columns = [i+'_non_anom' for i in group_non_basic_anom_obs.columns]
            group_basic_obs = group_basic_anom_obs.merge(group_non_basic_anom_obs, left_on = 'Cell', right_on = 'Cell')
            group_basic_obs = group_basic_obs[columns]

            detailled_output = pd.DataFrame()
            for cell in list(anom.Cell.unique()) :
                df_cell = group_basic_obs[group_basic_obs.index == cell]
                detailled_output = pd.concat([detailled_output, df_cell])
            cols = list(detailled_output.columns)
            new_cols = ['Cell']+cols
            detailled_output['Cell'] = detailled_output.index
            detailled_output = detailled_output[new_cols]
            for col in detailled_output.columns :
                if col != 'Cell' :
                    detailled_output[col] = round(detailled_output[col],3)

            new_cols = []
            new_cols.append('final_ranking')
            new_cols.append('Cell')
            new_cols.append('preclust')
            new_cols.append('last_anomaly')
            new_cols.append('last_anomaly_range')
            for col in detailled_output.columns :
                if col != 'Cell' :
                    new_cols.append(col)

            detailled_output = detailled_output.merge(main_output[['final_ranking','Cell','preclust','last_anomaly','last_anomaly_range']], left_on = detailled_output.index, right_on = 'Cell')
            detailled_output['last_anomaly'] = [str(i) for i in detailled_output['last_anomaly']]
            detailled_output['last_anomaly_range'] = [str(i) for i in detailled_output['last_anomaly_range']]
            detailled_output = detailled_output[new_cols]

            detailled_output_vide = pd.DataFrame()

            #df_detailled_output = detailled_output.values.tolist()

            detailled_output_table = dash_table.DataTable(
                id='detailled_output_table',
                columns=[{"name": i, "id": i} for i in detailled_output_vide.columns],
                data=detailled_output_vide.to_dict('records'),
                style_header={
                    'color': 'black',
                    'fontWeight': 'bold',
                    'fontSize' : '15px'
                    },
                style_cell={'textAlign': 'center'},
                sort_action="native",
                style_table={
                'height': 450,
                'overflowY': 'auto'
                },
                style_data = {'width':'auto'}
                )

            detailled_output.to_csv('data_dash/detailled_output.csv')

            detailled_output_table = dash_table.DataTable(
                id='detailled_output_table',
                columns=[{"name": i, "id": i} for i in detailled_output.columns],
                data=detailled_output.to_dict('records'),
                style_header={
                    'color': 'black',
                    'fontWeight': 'bold',
                    'fontSize' : '15px'
                    },
                style_cell={'textAlign': 'center'},
                sort_action="native",
                style_table={
                'height': 450,
                'overflowY': 'auto'
                },
                style_data = {'width':'auto'}
                )

            print('OK output exportation')

            anom_risk = anom

        ### INTERFERENCE CHARACTERIZATION ###
            df_data = pd.read_csv("data_dash/new_df.csv")
            df_data.index = [i for i in range(len(df_data))]
            list_anom_cells = list(anom.Cell)
            new_anom_risk = []
            for i in df_data.index:
                if df_data.loc[i, 'RTWP'] > max_RTWP:
                    new_anom_risk.append(1)
                else:
                    new_anom_risk.append(df_data.loc[i, 'Anom_risk'])
            df_data['new_anom_risk'] = new_anom_risk
            anom_neigh = pd.DataFrame(data={'Cell': list_anom_cells})

            anom_neigh = anom_neigh.merge(crown_data, left_on='Cell', right_on='Cell')

            crown = []
            for i in anom_neigh.index:
                liste = anom_neigh.loc[i, 'Crown']
                new_liste = []
                liste = liste.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' ', '')
                    new = new.replace("'", '')
                    new_liste.append(new)
                crown.append(new_liste)
            anom_neigh['Crown'] = crown

            anom_neighs = [];
            len_anom = []

            for i in anom_neigh.index:
                crown = anom_neigh.loc[i, 'Crown']
                neigh = []
                for cell in crown:
                    if cell in list_anom_cells:
                        neigh.append(cell)
                anom_neighs.append(neigh)
                len_anom.append(len(neigh))
            anom_neigh['Anom_neighbors'] = anom_neighs
            anom_neigh['len_anom'] = len_anom
            anom_neigh.to_csv('data_dash/anom_neigh.csv')

            df_data['Time'] = df_data['real_time']

            interf_chr = anom_neigh[anom_neigh.len_anom > 0]

            df_data = df_data[['Time', 'Cell', 'RTWP', 'Anom_risk', 'new_anom_risk', 'basic_anom']]

            ranges_total = []

            for j in interf_chr.index:
                ranges_begin = [];
                ranges_end = []
                df_site = df_data[df_data.Cell == interf_chr.loc[j, 'Cell']].sort_values(by='Time')
                for i in range(len(df_site)):
                    if i == 0:
                        if df_site.iloc[i, -1] == 1:
                            ranges_begin.append(list(df_site.Time)[i])

                    elif i == max(range(len(df_site))):
                        if df_site.iloc[i, -1] == 1:
                            ranges_end.append(list(df_site.Time)[i])

                    else:
                        if (df_site.iloc[i, -1] == 1) and (df_site.iloc[i - 1, -1] == 0):
                            ranges_begin.append(list(df_site.Time)[i])

                        if (df_site.iloc[i, -1] == 1) and (df_site.iloc[i + 1, -1] == 0):
                            ranges_end.append(list(df_site.Time)[i])

                ranges = []

                for k in range(len(ranges_begin)):
                    begin = ranges_begin[k];
                    end = ranges_end[k]
                    ranges.append([begin, end])

                ranges_total.append(ranges)

            interf_chr['Anom_ranges'] = ranges_total

            def get_delta(d1, d2):
                delta = d2 - d1
                return delta

            cells = [];
            neighbors = [];
            common_ranges = [];
            anom_ranges = []
            for i in interf_chr.index:
                cell = interf_chr.loc[i, 'Cell']
                anom_neighbors = interf_chr.loc[i, 'Anom_neighbors']
                anomaly_range = max(interf_chr[interf_chr.Cell == cell]['Anom_ranges'])

                for neigh in anom_neighbors:
                    if neigh in list(interf_chr.Cell.unique()):
                        neigh_anomaly_range = max(interf_chr[interf_chr.Cell == neigh]['Anom_ranges'])
                        for cell_rang in anomaly_range:
                            anom_time = []
                            cell_rang = [pd.to_datetime(cell_rang[0]), pd.to_datetime(cell_rang[1])]
                            delta = get_delta(cell_rang[0], cell_rang[1])
                            for i in range(delta.days * 24 + delta.seconds // 3600 + 1):
                                hour = cell_rang[0] + timedelta(hours=i)
                                anom_time.append(hour)

                            for neigh_rang in neigh_anomaly_range:
                                neigh_rang = [pd.to_datetime(neigh_rang[0]), pd.to_datetime(neigh_rang[1])]
                                delta = get_delta(neigh_rang[0], neigh_rang[1])
                                for i in range(delta.days * 24 + delta.seconds // 3600 + 1):
                                    hour = neigh_rang[0] + timedelta(hours=i)
                                    anom_time.append(hour)

                            unique_anom_time = list(set(anom_time))
                            common_range = 1 if len(anom_time) != len(unique_anom_time) else 0
                            cells.append(cell)
                            neighbors.append(neigh)
                            anom_ranges.append(cell_rang)
                            common_ranges.append(common_range)

            df_common_range = pd.DataFrame(
                data={'Cell': cells, 'Neighbor': neighbors, 'Anomaly_range': anom_ranges, 'Common_range': common_ranges})
            df_common_range.to_csv('data_dash/df_common_range.csv')

            df_data['Anom_risk'] = df_data['new_anom_risk']

            df_data = df_data[['Time', 'Cell', 'RTWP', 'basic_anom', 'Anom_risk', 'new_anom_risk']]

            risk_interval = []
            for i in df_common_range.index:
                cell = df_common_range.loc[i, 'Cell']
                anomaly_range = df_common_range.loc[i, 'Anomaly_range']
                anomaly_range = [str(anomaly_range[0]), str(anomaly_range[1])]

                df_cell = df_data[(df_data.Cell == cell) & (df_data.Time >= anomaly_range[0]) & (
                            df_data.Time <= anomaly_range[1])].sort_values(by='Time')
                df_cell.index = pd.to_datetime(df_cell.Time)

                if len(df_cell[df_cell.Anom_risk == 1]) > 0:
                    risk_interval.append(1)
                else:
                    risk_interval.append(0)
            df_common_range['Risk_interval'] = risk_interval

            new_risk_interval = []
            for i in df_common_range.index:
                cell = df_common_range.loc[i, 'Cell']
                anomaly_range = df_common_range.loc[i, 'Anomaly_range']
                new_rang = [str(anomaly_range[0] + timedelta(hours=-3)), str(anomaly_range[1] + timedelta(hours=3))]
                anomaly_range = new_rang

                df_cell = df_data[(df_data.Cell == cell) & (df_data.Time >= anomaly_range[0]) & (
                            df_data.Time <= anomaly_range[1])].sort_values(by='Time')
                df_cell.index = pd.to_datetime(df_cell.Time)

                if len(df_cell[df_cell.Anom_risk == 1]) > 0:
                    new_risk_interval.append(1)
                else:
                    new_risk_interval.append(0)
            df_common_range['New_risk_interval'] = new_risk_interval

            cells = [];
            neighbors = [];
            common_ranges = [];
            anom_ranges = []
            for i in interf_chr.index:
                cell = interf_chr.loc[i, 'Cell']
                anom_neighbors = interf_chr.loc[i, 'Anom_neighbors']
                anomaly_range = max(interf_chr[interf_chr.Cell == cell]['Anom_ranges'])

                for neigh in anom_neighbors:
                    if neigh in list(interf_chr.Cell.unique()):
                        neigh_anomaly_range = max(interf_chr[interf_chr.Cell == neigh]['Anom_ranges'])
                        for cell_rang in anomaly_range:
                            anom_time = []
                            cell_rang = [pd.to_datetime(cell_rang[0]), pd.to_datetime(cell_rang[1])]
                            delta = get_delta(cell_rang[0], cell_rang[1])
                            if str(delta) < '0 days 05:00:00':
                                for i in range(delta.days * 24 + delta.seconds // 3600 + 1):
                                    hour = cell_rang[0] + timedelta(hours=i)
                                    anom_time.append(hour)

                                for neigh_rang in neigh_anomaly_range:
                                    neigh_rang = [pd.to_datetime(neigh_rang[0]), pd.to_datetime(neigh_rang[1])]
                                    delta = get_delta(neigh_rang[0], neigh_rang[1])
                                    if str(delta) < '0 days 05:00:00':
                                        for i in range(delta.days * 24 + delta.seconds // 3600 + 1):
                                            hour = neigh_rang[0] + timedelta(hours=i)
                                            anom_time.append(hour)

                                unique_anom_time = list(set(anom_time))
                                common_range = 1 if len(anom_time) != len(unique_anom_time) else 0
                                cells.append(cell)
                                neighbors.append(neigh)
                                anom_ranges.append(cell_rang)
                                common_ranges.append(common_range)

            df_common_short_range = pd.DataFrame(
                data={'Cell': cells, 'Neighbor': neighbors, 'Anomaly_range': anom_ranges, 'Common_range': common_ranges})

            new_common_range = []
            for i in interf_chr.index:
                cell = interf_chr.loc[i, 'Cell']
                anom_neighbors = interf_chr.loc[i, 'Anom_neighbors']
                anomaly_range = max(interf_chr[interf_chr.Cell == cell]['Anom_ranges'])

                for neigh in anom_neighbors:
                    if neigh in list(interf_chr.Cell.unique()):
                        neigh_anomaly_range = max(interf_chr[interf_chr.Cell == neigh]['Anom_ranges'])
                        for cell_rang in anomaly_range:
                            anom_time = []
                            cell_rang = [pd.to_datetime(cell_rang[0]), pd.to_datetime(cell_rang[1])]
                            new_rang = [cell_rang[0] + timedelta(hours=-3), cell_rang[1] + timedelta(hours=3)]
                            # anomaly_range = new_rang
                            cell_rang = [pd.to_datetime(cell_rang[0]), pd.to_datetime(cell_rang[1])]
                            delta = get_delta(cell_rang[0], cell_rang[1])
                            delta_new = get_delta(new_rang[0], new_rang[1])
                            if str(delta) < '0 days 05:00:00':
                                for i in range(delta_new.days * 24 + delta_new.seconds // 3600 + 1):
                                    hour = new_rang[0] + timedelta(hours=i)
                                    anom_time.append(hour)

                                for neigh_rang in neigh_anomaly_range:
                                    neigh_rang = [pd.to_datetime(neigh_rang[0]), pd.to_datetime(neigh_rang[1])]
                                    new_rang = [neigh_rang[0] + timedelta(hours=-3), neigh_rang[1] + timedelta(hours=3)]
                                    neigh_rang = [pd.to_datetime(neigh_rang[0]), pd.to_datetime(neigh_rang[1])]
                                    delta = get_delta(neigh_rang[0], neigh_rang[1])
                                    delta_new = get_delta(new_rang[0], new_rang[1])
                                    if str(delta) < '0 days 05:00:00':
                                        for i in range(delta_new.days * 24 + delta_new.seconds // 3600 + 1):
                                            hour = new_rang[0] + timedelta(hours=i)
                                            anom_time.append(hour)

                                unique_anom_time = list(set(anom_time))
                                common_range = 1 if len(anom_time) != len(unique_anom_time) else 0
                                new_common_range.append(common_range)

            df_common_short_range['New_common_range'] = new_common_range

            common_anomaly = []
            for i in df_common_short_range.index:

                cell = df_common_short_range.loc[i, 'Cell']
                neigh = df_common_short_range.loc[i, 'Neighbor']
                rang = df_common_short_range.loc[i, 'Anomaly_range']

                rang = [str(rang[0]), str(rang[1])]

                df_cell = df_clean[df_clean.Cell == cell].sort_values(by='RTWP', ascending=False);
                df_cell = df_cell[(df_cell.Time >= rang[0]) & (df_cell.Time <= rang[1])]

                df_other = df_clean[df_clean.Cell == neigh].sort_values(by='RTWP', ascending=False);
                df_other = df_other[(df_other.Time >= rang[0]) & (df_other.Time <= rang[1])]

                if ((len(df_cell) > 0) and (len(df_other) > 0)):
                    if list(df_cell.Time)[0] == list(df_other.Time)[0]:
                        common_anomaly.append(1)
                    else:
                        common_anomaly.append(0)
                else:
                    common_anomaly.append('NaN')

            df_common_short_range['Common_anomaly'] = common_anomaly

            new_common_anomaly = []
            for i in df_common_short_range.index:

                cell = df_common_short_range.loc[i, 'Cell']
                neigh = df_common_short_range.loc[i, 'Neighbor']
                rang = df_common_short_range.loc[i, 'Anomaly_range']

                new_rang = [str(rang[0] + timedelta(hours=-3)), str(rang[1] + timedelta(hours=3))]
                rang = new_rang

                df_cell = df_clean[df_clean.Cell == cell].sort_values(by='RTWP', ascending=False);
                df_cell = df_cell[(df_cell.Time >= rang[0]) & (df_cell.Time <= rang[1])]

                df_other = df_clean[df_clean.Cell == neigh].sort_values(by='RTWP', ascending=False);
                df_other = df_other[(df_other.Time >= rang[0]) & (df_other.Time <= rang[1])]

                if ((len(df_cell) > 0) and (len(df_other) > 0)):
                    if list(df_cell.Time)[0] == list(df_other.Time)[0]:
                        new_common_anomaly.append(1)
                    else:
                        new_common_anomaly.append(0)
                else:
                    new_common_anomaly.append('NaN')

            df_common_short_range['New_common_anomaly'] = new_common_anomaly

            def cells_corr(df1, df2, method='spearman'):
                hour = [hour1 for hour1 in list(df1.Time) if hour1 in list(df2.Time)]
                df_cell1 = df1[df1.Time.isin(hour)];
                df_cell2 = df2[df2.Time.isin(hour)]
                if method == 'spearman':
                    return df_cell1, stats.spearmanr(df_cell1.RTWP, df_cell2.RTWP)[0]
                if method == 'kendall':
                    return df_cell1, stats.kendalltau(df_cell1.RTWP, df_cell2.RTWP)[0]

            corr = []

            for i in df_common_range.index:
                cell = df_common_range.loc[i, 'Cell']
                neigh = df_common_range.loc[i, 'Neighbor']
                rang = df_common_range.loc[i, 'Anomaly_range']

                rang = [str(rang[0]), str(rang[1])]

                df_cell = df_clean[df_clean.Cell == cell].sort_values(by='Time');
                df_cell = df_cell[(df_cell.Time >= rang[0]) & (df_cell.Time <= rang[1])]

                df_other = df_clean[df_clean.Cell == neigh].sort_values(by='Time');
                df_other = df_other[(df_other.Time >= rang[0]) & (df_other.Time <= rang[1])]
                df_1, correlation = cells_corr(df_cell, df_other)
                if len(df_1) > 4:
                    corr.append(correlation)
                else:
                    corr.append('NaN')

            df_common_range['Corr'] = corr

            df_common_short_range = df_common_short_range[
                ['Cell', 'Neighbor', 'Anomaly_range', 'New_common_range', 'New_common_anomaly']]
            df_common_short_range.columns = ['Cell', 'Neighbor', 'Anomaly_range', 'Common_range', 'Common_anomaly']

            df_short_no_trend = df_common_short_range[df_common_short_range.Common_anomaly.isin([0,'0','NaN'])]

            df_short_no_trend = df_short_no_trend[['Cell','Neighbor','Anomaly_range','Common_range']]

            df_short = df_common_short_range[(df_common_short_range.Common_range == 1) & (df_common_short_range.Common_anomaly == 1)]

            df_no_trend = df_common_range[(df_common_range.Common_range == 0) | (df_common_range.Corr == 'NaN')]

            df_trend = df_common_range[(df_common_range.Common_range == 1) & (
                        df_common_range.Corr != 'NaN')]  # & (df_common_range.New_risk_interval == 1)]

            df_no_trend_other = df_trend[df_trend.Corr <= 0.85]

            df_no_external = pd.concat([df_short_no_trend, df_no_trend, df_no_trend_other])

            df_no_external['Anomaly_range_str'] = [str(i) for i in df_no_external.Anomaly_range]

            df_no_external.to_csv('data_dash/df_no_external_bef.csv')

            df_trend = df_trend[df_trend.Corr > 0.85]

            df_trend['Common_anomaly'] = [1 if i > 0.85 else 0 for i in df_trend.Corr]
            df_long = df_trend[['Cell', 'Neighbor', 'Anomaly_range', 'Common_range', 'Common_anomaly']]
            df_long.columns = ['Cell', 'Neighbor', 'Anomaly_range', 'Common_range', 'Common_anomaly']

            df_all = pd.concat([df_long, df_short])

            df_all['Anomaly_range_str'] = [str(i) for i in df_all.Anomaly_range]

            df_all.index = list(range(len(df_all)))

            df_no_external = df_no_external.merge(df_all[['Cell','Anomaly_range_str','Common_anomaly']], how = 'left', left_on = ['Cell','Anomaly_range_str'], right_on = ['Cell', 'Anomaly_range_str'])
            df_no_final = df_no_external[~df_no_external.Common_anomaly.isin([1,'1.0','1'])]

            df_no_final = df_no_final[['Cell','Anomaly_range_str']].drop_duplicates()

            anomaly_range = []
            for i in df_no_final.index :
                date = df_no_final.loc[i, 'Anomaly_range_str']
                new_date = []
                date = date.split(',')
                for new in date:
                    new = new.replace('Timestamp(', '')
                    new = new.replace(')', '')
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new_date.append(pd.to_datetime(new))
                anomaly_range.append(new_date)
            df_no_final['Anomaly_range'] = anomaly_range

            df_no_final = df_no_final[['Cell','Anomaly_range']]
            df_data.to_csv('data_dash/df_data.csv')
            new_df = pd.read_csv('data_dash/new_df.csv')
            #new_df = df_data
            since_anom_end = [] ; anom_duration = [] ; qos_degradation = [] ; maximum_RTWP = []
            for i in df_no_final.index :
                cell = df_no_final.loc[i,'Cell']
                anom_range = df_no_final.loc[i, 'Anomaly_range']
                since_anom_end.append(pd.to_datetime(max(new_df[new_df.Cell == cell]['real_time']))-anom_range[1])
                anom_duration.append(anom_range[1]-anom_range[0])

                df_anom = new_df[((new_df.Cell == cell) & (pd.to_datetime(new_df.real_time) >= anom_range[0]) & (
                            pd.to_datetime(new_df.real_time) <= anom_range[1]))]
                df_normal = new_df[((new_df.Cell == cell) & (new_df.basic_anom == 0))]

                maximum_RTWP.append(max(df_anom.RTWP))
                qos = [];
                poids = 0

                if (('DCR-PS' in new_df.columns) & ('DCR-PS' in qos_kpi)):
                    dcr = (df_anom['DCR-PS'].mean() - df_normal['DCR-PS'].mean()) * poids_DCRPS
                    poids += poids_DCRPS
                    qos.append(dcr)
                else:
                    dcr = 0
                if (('CSSR-PS' in new_df.columns) & ('CSSR-PS' in qos_kpi)):
                    cssrps = (df_normal['CSSR-PS'].mean() - df_anom['CSSR-PS'].mean()) * poids_CSSRPS
                    poids += poids_CSSRPS
                    qos.append(cssrps)
                else:
                    cssrps = 1
                if (('CSSR-CS' in new_df.columns) & ('CSSR-CS' in qos_kpi)):
                    cssrcs = (df_normal['CSSR-CS'].mean() - df_anom['CSSR-CS'].mean()) * poids_CSSRCS
                    poids += poids_CSSRCS
                    qos.append(cssrcs)
                else:
                    cssrcs = 1
                if len(qos) > 0:
                    qos_degradation.append(round(sum(qos) / poids, 3))
                else:
                    qos_degradation.append('NaN')

            df_no_final['Since_anom_end'] = since_anom_end
            df_no_final['Anom_duration'] = anom_duration
            df_no_final['qos_degradation'] = qos_degradation
            df_no_final['max_RTWP'] = maximum_RTWP

            df_no_final['ranking'] = ''

            for i in df_no_final.index:
                if 'max RTWP' in indicators :
                    if df_no_final.loc[i, 'max_RTWP'] > -85:
                        df_no_final['ranking'][i] += '1'
                    elif df_no_final.loc[i, 'max_RTWP'] > -90:
                        df_no_final['ranking'][i] += '2'
                    if df_no_final.loc[i, 'max_RTWP'] > -95:
                        df_no_final['ranking'][i] += '3'
                    if df_no_final.loc[i, 'max_RTWP'] > -100:
                        df_no_final['ranking'][i] += '4'
                    else:
                        df_no_final['ranking'][i] += '5'

                if 'last anomaly' in indicators :
                    if df_no_final.loc[i, 'Since_anom_end'] < timedelta(days=3):
                        df_no_final['ranking'][i] += '1'
                    elif df_no_final.loc[i, 'Since_anom_end'] < timedelta(days=7):
                        df_no_final['ranking'][i] += '2'
                    else:
                        df_no_final['ranking'][i] += '3'

                if 'anomaly duration' in indicators :
                    if int(str(df_no_final.loc[i, 'Anom_duration']).split(' ')[0]) > 2:
                        df_no_final['ranking'][i] += '1'
                    else:
                        df_no_final['ranking'][i] += '2'

            if 'qos degradation' in indicators :
                df_no_final = df_no_final.sort_values(by=['ranking', 'qos_degradation'], ascending=[True, False])
            else :
                df_no_final = df_no_final.sort_values(by='ranking', ascending=True)

            i = 0
            df_no_final['final_ranking'] = 0
            for line in df_no_final.index:
                i += 1
                df_no_final['final_ranking'][line] = i
            df_no_final.index = [i for i in range(len(df_no_final))]

            for col in df_no_final.columns :
                df_no_final[col] = [str(i) for i in df_no_final[col]]
            df_no_final = df_no_final.drop_duplicates()

            df_no_external.to_csv('data_dash/df_no_external_aft.csv')
            df_no_final.to_csv('data_dash/df_no_final.csv')

            ext_interf = pd.DataFrame(
                df_all[['Cell', 'Anomaly_range_str', 'Neighbor']].groupby(['Cell', 'Anomaly_range_str'])
                ['Neighbor'].apply(list))

            anomaly_range = []
            for i in range(len(ext_interf)):
                date = ext_interf.index[i][1]
                new_date = []
                date = date.split(',')
                for new in date:
                    new = new.replace('Timestamp(', '')
                    new = new.replace(')', '')
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new_date.append(pd.to_datetime(new))
                anomaly_range.append(new_date)
            ext_interf['Anomaly_range'] = anomaly_range

            ext_interf['Anom_end'] = [str(i[1]) for i in ext_interf.Anomaly_range]

            final_impacted_cells = []
            for i in range(len(ext_interf)):
                impacted_cells = list(ext_interf.Neighbor[i])
                impacted_cells.append(ext_interf.index[i][0])
                final_impacted_cells.append(impacted_cells)
            ext_interf['Neighbor'] = final_impacted_cells

            ext_interf['Number_of_impacted_cells'] = [len(i) for i in ext_interf.Neighbor]

            ext_interf.index = list(range(len(ext_interf)))

            neighbor_loc = []
            for i in ext_interf.index:
                neighbors = ext_interf.loc[i, 'Neighbor']
                neigh_local = []
                for neigh in neighbors:
                    localisation = [max(df_loc[df_loc.Cell == neigh]['Longitude']),
                                    max(df_loc[df_loc.Cell == neigh]['Latitude'])]
                    neigh_local.append(localisation)
                neighbor_loc.append(neigh_local)
            ext_interf['Neighbor_loc'] = neighbor_loc

            ext_interf['Since_anom_end'] = [pd.to_datetime(max(df_clean.Time)) - pd.to_datetime(i) for i in
                                            ext_interf.Anom_end]

            ext_interf['Anom_duration'] = [pd.to_datetime(i[1]) - pd.to_datetime(i[0]) for i in ext_interf['Anomaly_range']]

            qos_degradation = [] ; maximum_RTWP = []
            for i in ext_interf.index:
                qos = [];
                poids = 0
                max_rtwp_cell = -2000
                for cell in ext_interf.loc[i, 'Neighbor']:
                    df_cell = df_clean[
                        (df_clean.Cell == cell) & (df_clean.Time >= str(ext_interf.loc[i, 'Anomaly_range'][0])) & (
                                df_clean.Time <= str(ext_interf.loc[i, 'Anomaly_range'][1]))]

                    df_normal = new_df[((new_df.Cell == cell) & (new_df.basic_anom == 0))]

                    if len(df_cell) > 0:

                        if max(df_cell.RTWP) > max_rtwp_cell:
                            max_rtwp_cell = max(df_cell.RTWP)

                        if (('DCR-PS' in df_cell.columns) & ('DCR-PS' in qos_kpi)):
                            dcr = (df_cell['DCR-PS'].mean() - df_normal['DCR-PS'].mean()) * poids_DCRPS
                            poids += poids_DCRPS
                            qos.append(dcr)
                        else:
                            dcr = 0
                        if (('CSSR-PS' in df_cell.columns) & ('CSSR-PS' in qos_kpi)):
                            cssrps = (df_normal['CSSR-PS'].mean() - df_cell['CSSR-PS'].mean()) * poids_CSSRPS
                            poids += poids_CSSRPS
                            qos.append(cssrps)
                        else:
                            cssrps = 1
                        if (('CSSR-CS' in df_cell.columns) & ('CSSR-CS' in qos_kpi)):
                            cssrcs = (df_normal['CSSR-CS'].mean() - df_cell['CSSR-CS'].mean()) * poids_CSSRCS
                            poids += poids_CSSRCS
                            qos.append(cssrcs)
                        else:
                            cssrcs = 1

                if len(qos) > 0:
                    qos_degradation.append(round(sum(qos) / poids, 3))
                else:
                    qos_degradation.append('NaN')
                maximum_RTWP.append(max_rtwp_cell)
            ext_interf['qos_degradation'] = qos_degradation
            ext_interf['max_RTWP'] = maximum_RTWP

            ext_interf['ranking'] = ''

            for i in ext_interf.index:
                if 'max RTWP' in indicators:
                    if ext_interf.loc[i, 'max_RTWP'] > -85:
                        ext_interf['ranking'][i] += '1'
                    elif ext_interf.loc[i, 'max_RTWP'] > -90:
                        ext_interf['ranking'][i] += '2'
                    if ext_interf.loc[i, 'max_RTWP'] > -95:
                        ext_interf['ranking'][i] += '3'
                    if ext_interf.loc[i, 'max_RTWP'] > -100:
                        ext_interf['ranking'][i] += '4'
                    else:
                        ext_interf['ranking'][i] += '5'

                if 'last anomaly' in indicators:
                    if ext_interf.loc[i, 'Since_anom_end'] < timedelta(days=3):
                        ext_interf['ranking'][i] += '1'
                    elif ext_interf.loc[i, 'Since_anom_end'] < timedelta(days=7):
                        ext_interf['ranking'][i] += '2'
                    else:
                        ext_interf['ranking'][i] += '3'

                if 'anomaly duration' in indicators:
                    if int(str(df_no_final.loc[i, 'Anom_duration']).split(' ')[0]) > 2:
                        ext_interf['ranking'][i] += '1'
                    else:
                        ext_interf['ranking'][i] += '2'

            if 'qos degradation' in indicators:
                ext_interf = ext_interf.sort_values(by=['ranking', 'qos_degradation'], ascending=[True, False])
            else:
                ext_interf = ext_interf.sort_values(by='ranking', ascending=True)

            i = 0
            ext_interf['final_ranking'] = 0
            for line in ext_interf.index:
                i += 1
                ext_interf['final_ranking'][line] = i
            ext_interf.index = [i for i in range(len(ext_interf))]

            ext_interf_prio = ext_interf[
                ['Neighbor', 'Since_anom_end', 'Anom_duration', 'qos_degradation', 'final_ranking']]
            ext_interf_prio.columns = ['Cells', 'Since_anom_end', 'Anom_duration', 'qos_degradation', 'final_ranking']
            ext_interf_prio['Since_anom_end'] = [str(i) for i in ext_interf_prio['Since_anom_end']]
            ext_interf_prio['Anom_duration'] = [str(i) for i in ext_interf_prio['Anom_duration']]
            #ext_interf_prio['QoS_degradation'] = [round(i, 3) for i in ext_interf_prio['QoS_degradation']]



            ext_interf.to_csv('data_dash/ext_interf.csv')
            ext_interf_prio.to_csv('data_dash/ext_interf_prio.csv')

            ext_interf_prio['Cells_str'] = [str(i) for i in ext_interf_prio.Cells
                                            ]
            n_clicks = 0

    ###### INTERFERENCE LOCALIZATION ########
            def get_intersections(circle_1, circle_2):
                # circle 1: (x0, y0), radius r0
                # circle 2: (x1, y1), radius r1
                x0 = circle_1[0][0]
                y0 = circle_1[0][1]
                r0 = circle_1[1]
                x1 = circle_2[0][0]
                y1 = circle_2[0][1]
                r1 = circle_2[1]
                d = math.sqrt((x1 - x0) ** 2 + (y1 - y0) ** 2)

                # non intersecting
                if d > r0 + r1:
                    return None
                # One circle within other
                if d < abs(r0 - r1):
                    return None
                # coincident circles
                if d == 0 and r0 == r1:
                    return None
                else:
                    a = (r0 ** 2 - r1 ** 2 + d ** 2) / (2 * d)
                    h = math.sqrt(r0 ** 2 - a ** 2)
                    x2 = x0 + a * (x1 - x0) / d
                    y2 = y0 + a * (y1 - y0) / d
                    x3 = x2 + h * (y1 - y0) / d
                    y3 = y2 - h * (x1 - x0) / d

                    x4 = x2 - h * (y1 - y0) / d
                    y4 = y2 + h * (x1 - x0) / d

                    return (x3, y3, x4, y4)

            def circle_line_intersection(circle_center, circle_radius, pt1, pt2, full_line=True, tangent_tol=1e-9):
                """ Find the points at which a circle intersects a line-segment.  This can happen at 0, 1, or 2 points.

                :param circle_center: The (x, y) location of the circle center
                :param circle_radius: The radius of the circle
                :param pt1: The (x, y) location of the first point of the segment
                :param pt2: The (x, y) location of the second point of the segment
                :param full_line: True to find intersections along full line - not just in the segment.  False will just return intersections within the segment.
                :param tangent_tol: Numerical tolerance at which we decide the intersections are close enough to consider it a tangent
                :return Sequence[Tuple[float, float]]: A list of length 0, 1, or 2, where each element is a point at which the circle intercepts a line segment.

                Note: We follow: http://mathworld.wolfram.com/Circle-LineIntersection.html
                """

                (p1x, p1y), (p2x, p2y), (cx, cy) = pt1, pt2, circle_center
                (x1, y1), (x2, y2) = (p1x - cx, p1y - cy), (p2x - cx, p2y - cy)
                dx, dy = (x2 - x1), (y2 - y1)
                dr = (dx ** 2 + dy ** 2) ** .5
                big_d = x1 * y2 - x2 * y1
                discriminant = circle_radius ** 2 * dr ** 2 - big_d ** 2

                if discriminant < 0:  # No intersection between circle and line
                    return []
                else:  # There may be 0, 1, or 2 intersections with the segment
                    intersections = [
                        (cx + (big_d * dy + sign * (-1 if dy < 0 else 1) * dx * discriminant ** .5) / dr ** 2,
                         cy + (-big_d * dx + sign * abs(dy) * discriminant ** .5) / dr ** 2)
                        for sign in
                        ((1, -1) if dy < 0 else (-1, 1))]  # This makes sure the order along the segment is correct
                    if not full_line:  # If only considering the segment, filter out intersections that do not fall within the segment
                        fraction_along_segment = [(xi - p1x) / dx if abs(dx) > abs(dy) else (yi - p1y) / dy for xi, yi in
                                                  intersections]
                        intersections = [pt for pt, frac in zip(intersections, fraction_along_segment) if 0 <= frac <= 1]
                    if len(intersections) == 2 and abs(
                            discriminant) <= tangent_tol:  # If line is tangent to circle, return just one point (as both intersections have same location)
                        return [intersections[0]]
                    else:
                        return intersections

            new_neighbor_loc = [];
            #R = 6371.009
            df_cart = pd.DataFrame();
            cells = [];
            x_cartesien = [];
            y_cartesien = [];
            #z_cartesien = []
            for i in ext_interf.index:
                new_loc = []
                for neigh in ext_interf.loc[i, 'Neighbor']:
                    cells.append(neigh)
                for loc in ext_interf.loc[i, 'Neighbor_loc']:
                    # loc = [loc[0]-origin_lon,loc[1]-origin_lat]
                    x_cart = loc[0]
                    y_cart = loc[1]
                    # z_cart = R * np.sin(loc[1])
                    x_cartesien.append(x_cart)
                    y_cartesien.append(y_cart)
                    # z_cartesien.append(z_cart)
                    new_loc.append([x_cart, y_cart])
                new_neighbor_loc.append(new_loc)
            ext_interf['Neighbor_loc_cart'] = new_neighbor_loc

            df_cart['Cell'] = cells;
            df_cart['x_cart'] = x_cartesien;
            df_cart['y_cart'] = y_cartesien;
            # df_cart['z_cart'] = z_cartesien
            #
            # origin_x = df_cart.loc[0, 'x_cart']
            # origin_y = df_cart.loc[0, 'y_cart']
            # origin_z = df_cart.loc[0, 'z_cart']
            #
            # for i in df_cart.index:
            #     df_cart.loc[i, 'x_cart'] = df_cart.loc[i, 'x_cart'] - origin_x
            #     df_cart.loc[i, 'y_cart'] = df_cart.loc[i, 'y_cart'] - origin_y
            #     df_cart.loc[i, 'z_cart'] = df_cart.loc[i, 'z_cart'] - origin_z

            circles = [];
            intersections = []
            for i in ext_interf.index:

                neighbor_pairs = [list(x) for x in itertools.combinations(ext_interf.loc[i, 'Neighbor'], 2)]

                barycentre_coord = []
                circles_data = []
                intersec = []
                cells_pairs = []
                for pair in neighbor_pairs:
                    if pair[0].split('_')[:-1] != pair[1].split('_')[:-1]:
                        cells_pairs.append(pair)
                        anomaly_range = ext_interf.loc[i, 'Anomaly_range']
                        if str(pd.to_datetime(anomaly_range[1]) - pd.to_datetime(anomaly_range[0])) < '0 days 05':
                            anomaly_range = [anomaly_range[0] + timedelta(hours=-3), anomaly_range[1] + timedelta(hours=3)]

                        rtwp_max_1 = max(df_clean[(df_clean.Cell == pair[0]) & (df_clean.Time >= str(anomaly_range[0])) & (
                                    df_clean.Time <= str(anomaly_range[1]))]['RTWP'])
                        rtwp_max_2 = max(df_clean[(df_clean.Cell == pair[1]) & (df_clean.Time >= str(anomaly_range[0])) & (
                                    df_clean.Time <= str(anomaly_range[1]))]['RTWP'])

                        rtwp_max_1 = rtwp_max_1 / 10
                        rtwp_max_2 = rtwp_max_2 / 10

                        alpha = math.exp(rtwp_max_1) / math.exp(rtwp_max_2)

                        x_barycentre_i = (max(df_cart[df_cart.Cell == pair[0]]['x_cart']) + alpha * max(
                            df_cart[df_cart.Cell == pair[1]]['x_cart'])) / (1 + alpha)
                        y_barycentre_i = (max(df_cart[df_cart.Cell == pair[0]]['y_cart']) + alpha * max(
                            df_cart[df_cart.Cell == pair[1]]['y_cart'])) / (1 + alpha)
                        barycentre_i = [x_barycentre_i, y_barycentre_i]
                        x_barycentre_j = (max(df_cart[df_cart.Cell == pair[0]]['x_cart']) - alpha * max(
                            df_cart[df_cart.Cell == pair[1]]['x_cart'])) / (1 - alpha)
                        y_barycentre_j = (max(df_cart[df_cart.Cell == pair[0]]['y_cart']) - alpha * max(
                            df_cart[df_cart.Cell == pair[1]]['y_cart'])) / (1 - alpha)
                        barycentre_j = [x_barycentre_j, y_barycentre_j]

                        middle_circle = [(barycentre_i[0] + barycentre_j[0]) / 2, (barycentre_i[1] + barycentre_j[1]) / 2]
                        barycentre_coord.append([barycentre_i, barycentre_j])

                        dist_ij = math.sqrt(
                            (barycentre_i[0] - barycentre_j[0]) ** 2 + (barycentre_i[1] - barycentre_j[1]) ** 2)
                        circles_data.append([middle_circle, dist_ij / 2])
                indice = 0
                if len(circles_data) > 1:
                    circle_pairs = [list(x) for x in itertools.combinations(circles_data, 2)]
                    cells = cells_pairs[indice]
                    indice += 1
                    for pair in circle_pairs:
                        val = get_intersections(pair[0], pair[1])
                        if val == None:
                            test0 = circle_line_intersection(pair[0][0], pair[0][1], (
                            max(df_cart[df_cart.Cell == cells[0]]['x_cart']),
                            max(df_cart[df_cart.Cell == cells[0]]['y_cart'])), (
                                                             max(df_cart[df_cart.Cell == cells[1]]['x_cart']),
                                                             max(df_cart[df_cart.Cell == cells[1]]['y_cart'])))
                            test1 = circle_line_intersection(pair[1][0], pair[1][1], (
                            max(df_cart[df_cart.Cell == cells[0]]['x_cart']),
                            max(df_cart[df_cart.Cell == cells[0]]['y_cart'])), (
                                                             max(df_cart[df_cart.Cell == cells[1]]['x_cart']),
                                                             max(df_cart[df_cart.Cell == cells[1]]['y_cart'])))
                            val = [(max(df_cart[df_cart.Cell == cells[0]]['x_cart']) * pair[0][1] + max(
                                df_cart[df_cart.Cell == cells[1]]['x_cart']) * pair[1][1]) / (pair[0][1] + pair[1][1]), (
                                               max(df_cart[df_cart.Cell == cells[0]]['y_cart']) * pair[0][1] + max(
                                           df_cart[df_cart.Cell == cells[1]]['y_cart']) * pair[1][1]) / (
                                               pair[0][1] + pair[1][1])]
                            dist_final0 = []
                            for pt in test0:
                                dist = math.sqrt((pt[0] - val[0]) ** 2 + (pt[1] - val[1]) ** 2)
                                dist_final0.append(dist)
                            dist_final1 = []
                            for pt in test1:
                                dist = math.sqrt((pt[0] - val[0]) ** 2 + (pt[1] - val[1]) ** 2)
                                dist_final1.append(dist)
                            if len(dist_final0) > 0:
                                test0 = test0[dist_final0.index(min(dist_final0))]
                            else:
                                val = []
                            if len(dist_final1) > 0:
                                test1 = test1[dist_final1.index(min(dist_final1))]
                            else:
                                val = []

                            if (len(dist_final0) > 0) & (len(dist_final1) > 0):
                                val = [(test0[0] * pair[0][1] + test1[0] * pair[1][1]) / (pair[0][1] + pair[1][1]),
                                       (test0[1] * pair[0][1] + test1[1] * pair[1][1]) / (pair[0][1] + pair[1][1])]

                        intersec.append(val)

                circles.append(circles_data)

                #         if len(intersec) == 0 :
                #             intersec.append(["Circle with ({},{}) center and {} radius".format()

                intersections.append(intersec)

            ext_interf['Circles'] = circles
            ext_interf['Intersections'] = intersections
            ext_interf_loc = ext_interf[['Neighbor','Intersections']]
            ext_interf_loc.columns = ['Cells','Localization']
            ext_interf_loc['Cells'] = [str(i) for i in ext_interf_loc.Cells]
            ext_interf_loc['Localization'] = [str(i) for i in ext_interf_loc.Localization]
            #print(ext_interf_loc)
            ext_interf['Cells'] = [str(i) for i in ext_interf.Neighbor]
            ext_interf = ext_interf.merge(ext_interf_loc[['Cells','Localization']], left_on = 'Cells', right_on = 'Cells')
            for col in ext_interf.columns:
                ext_interf[col] = [str(i) for i in ext_interf[col]]

            ext_interf = ext_interf.drop_duplicates()

            ext_interf.to_csv('data_dash/ext_interf.csv')

            return(dcc.Graph(figure = fig), output_table_pc, detailled_output_table,
                   html.H6('Detection course ended successfully.', style={'color': 'green'}),
                   html.Button(id='excel_button', children='Export Excel', n_clicks=0)
                   )
        except:
            return(None, None, None,
                   html.H6('Error during the detection course.', style={'color': 'red'}),
                   None)
    
    else :
        return(None, None, None, None, None)


    
@app.callback([Output('dropdown_cell', 'children'),
               Output('interf_table', 'children')],
 			[Input('graph_button_3', 'n_clicks'),
              Input('radio_pim','value')])

def dropdown_pim(n_clicks, choice):
    if n_clicks >= 1 :

        df_data = pd.read_csv('data_dash/new_df.csv')

        #anom = pd.read_csv('data_dash/anom.csv')
        df_no_final = pd.read_csv('data_dash/df_no_final.csv')
        ext_interf = pd.read_csv('data_dash/ext_interf.csv')

        if choice == 'PIM' :
            table_cells = df_no_final
            del table_cells['Unnamed: 0']
            del table_cells['ranking']
            table_table = table_cells[['Cell', 'Anomaly_range', 'Since_anom_end', 'Anom_duration', 'qos_degradation', 'max_RTWP', 'final_ranking']]
            new_name = [] ; indice = 1
            for i in table_cells.index :
                new_name.append(str(i)+table_cells.loc[i, 'Cell'])
            table_cells['Cell_dropdown'] = new_name
            table_table = table_cells
            table_cells.to_csv('data_dash/table_cells.csv')

            table = dash_table.DataTable(
                id='table',
                columns=[{"name": i, "id": i} for i in table_table[
                    ['Cell', 'Since_anom_end', 'Anom_duration', 'qos_degradation', 'max_RTWP',
                     'final_ranking']].columns],
                data=table_table[['Cell', 'Since_anom_end', 'Anom_duration', 'qos_degradation', 'max_RTWP',
                                  'final_ranking']].to_dict('records'),
                style_header={
                    'color': 'black',
                    'fontWeight': 'bold',
                    'fontSize': '15px'
                },
                style_cell={'textAlign': 'center'},
                sort_action="native",
                style_table={
                    'height': 450,
                    'overflowY': 'auto'
                },
                style_data={'width': 'auto'}
            )

        else :
            table_cells = ext_interf
            table_cells['Cells'] = table_cells['Neighbor']
            table_cells['Cell'] = table_cells['Cells']
            table_cells = table_cells[['Cell','Anomaly_range','Since_anom_end','Anom_duration','qos_degradation','max_RTWP','Localization','final_ranking']]
            new_name = [] ; indice = 1
            for i in table_cells.index :
                new_name.append(str(i)+table_cells.loc[i, 'Cell'])
            table_cells['Cell_dropdown'] = new_name
            table_table = table_cells
            table_cells.to_csv('data_dash/table_cells.csv')

            table = dash_table.DataTable(
                id='table',
                columns=[{"name": i, "id": i} for i in table_table[['Cell','Since_anom_end','Anom_duration','qos_degradation','max_RTWP','Localization','final_ranking']].columns],
                data=table_table[['Cell','Since_anom_end','Anom_duration','qos_degradation','max_RTWP','Localization','final_ranking']].to_dict('records'),
                style_header={
                    'color': 'black',
                    'fontWeight': 'bold',
                    'fontSize': '15px'
                },
                style_cell={'textAlign': 'center'},
                sort_action="native",
                style_table={
                    'height': 450,
                    'overflowY': 'auto'
                },
                style_data={'width': 'auto'}
            )

        return(dcc.Dropdown(
                            options=[{'label' : cell, 'value' : cell} for cell in table_cells.Cell_dropdown.unique()],
                            searchable = False,
                                  placeholder = 'Choose a cell...',
                                  style = {'width' : '50%'},
                                  value = table_cells.Cell_dropdown.unique()[0]
                              ),
        table)

    else :
        return(None, None)

@app.callback(Output('interf_graph', 'children'),
 			[Input('graph_button_pim', 'n_clicks'),
             Input('dropdown_cell','children'),
             Input('radio_pim','value')])

def anomaly_graph(n_clicks, cell, choice):
    if n_clicks >= 1 :
        try:
            cell = cell['props']['value']
            df = pd.read_csv('data_dash/df_clean.csv')

            table_cells = pd.read_csv('data_dash/table_cells.csv')
            anomaly = max(table_cells[table_cells.Cell_dropdown == cell]['Anomaly_range'])

            cell = max(table_cells[table_cells.Cell_dropdown == cell].Cell)

            new_liste = []
            liste = cell.split(',')
            for new in liste:
                new = new.replace('[', '')
                new = new.replace(']', '')
                new = new.replace(' ', '')
                new = new.replace("'", '')
                new_liste.append(new)
            cells = new_liste
            df_cells = pd.DataFrame()
            for cell in cells :
                df_cell = df[df.Cell == cell].sort_index()
                df_cells = pd.concat([df_cells, df_cell])

            if choice == 'PIM' :
                fig = px.line(df_cells, x = 'Time', y = 'RTWP')
            else :
                fig = px.line(df_cells, x = 'Time', y = 'RTWP', color = 'Cell')

            new_liste = []
            liste = anomaly.split(',')
            for new in liste:
                new = new.replace('[', '')
                new = new.replace(']', '')
                new = new.replace(' Timestamp(', '')
                new = new.replace('Timestamp(', '')
                new = new.replace(')', '')
                new = new.replace("'", '')
                new_liste.append(pd.to_datetime(new))
            anomaly = new_liste

            fig.add_shape(type="rect",
                          x0=anomaly[0], y0=min(df_cells.RTWP) - 1, x1=anomaly[1], y1=max(df_cells.RTWP) + 5,
                          line=dict(
                              color="Crimson",
                              width=2,
                          ),
                          )
            fig.update_yaxes(range=[min(df_cell.RTWP)-1, max(df_cell.RTWP)+5])

            n_clicks = 0
            return(dcc.Graph(figure = fig))
        except:
            return(None)

 
     
@app.callback(Output('kpi_hsupa', 'children'),
 			[Input('graph_button_4', 'n_clicks'),
              Input('dropdown_cell','children'),
             Input('radio_pim','value')])
              
def kpi_hsupa(n_clicks, cell, choice) :
    if n_clicks >= 1 :

        try:
            cell = cell['props']['value']
            df = pd.read_csv('data_dash/df_clean.csv')

            table_cells = pd.read_csv('data_dash/table_cells.csv')
            anomaly = max(table_cells[table_cells.Cell_dropdown == cell]['Anomaly_range'])

            cell = max(table_cells[table_cells.Cell_dropdown == cell].Cell)

            if 'Avg Num HSUPA Users' in df.columns :

                new_liste = []
                liste = cell.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' ', '')
                    new = new.replace("'", '')
                    new_liste.append(new)
                cells = new_liste
                df_cells = pd.DataFrame()
                for cell in cells:
                    df_cell = df[df.Cell == cell].sort_index()
                    df_cells = pd.concat([df_cells, df_cell])

                if choice == 'PIM':
                    fig = px.line(df_cells, x='Time', y='Avg Num HSUPA Users')
                else:
                    fig = px.line(df_cells, x='Time', y='Avg Num HSUPA Users', color='Cell')

                new_liste = []
                liste = anomaly.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' Timestamp(', '')
                    new = new.replace('Timestamp(', '')
                    new = new.replace(')', '')
                    new = new.replace("'", '')
                    new_liste.append(pd.to_datetime(new))
                anomaly = new_liste

                fig.add_shape(type="rect",
                                      x0=anomaly[0], y0=min(df_cells['Avg Num HSUPA Users']), x1=anomaly[1], y1=max(df_cells['Avg Num HSUPA Users']),
                                      line=dict(
                                          color="Crimson",
                                          width=2,
                                      ),
                                      )

                return(dcc.Graph(figure = fig))

            else :
                return(None)
        except:
            return(None)

    
@app.callback(Output('kpi_hsdpa', 'children'),
 			[Input('graph_button_4', 'n_clicks'),
             Input('dropdown_cell', 'children'),
             Input('radio_pim', 'value')])
def kpi_hsdpa(n_clicks, cell, choice):
    if n_clicks >= 1 :

        try:
            cell = cell['props']['value']
            df = pd.read_csv('data_dash/df_clean.csv')

            table_cells = pd.read_csv('data_dash/table_cells.csv')
            anomaly = max(table_cells[table_cells.Cell_dropdown == cell]['Anomaly_range'])
            cell = max(table_cells[table_cells.Cell_dropdown == cell].Cell)

            if 'Avg Num HSDPA Users' in df.columns:

                new_liste = []
                liste = cell.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' ', '')
                    new = new.replace("'", '')
                    new_liste.append(new)
                cells = new_liste
                df_cells = pd.DataFrame()
                for cell in cells:
                    df_cell = df[df.Cell == cell].sort_index()
                    df_cells = pd.concat([df_cells, df_cell])

                if choice == 'PIM':
                    fig = px.line(df_cells, x='Time', y='Avg Num HSDPA Users')
                else:
                    fig = px.line(df_cells, x='Time', y='Avg Num HSDPA Users', color='Cell')

                new_liste = []
                liste = anomaly.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' Timestamp(', '')
                    new = new.replace('Timestamp(', '')
                    new = new.replace(')', '')
                    new = new.replace("'", '')
                    new_liste.append(pd.to_datetime(new))
                anomaly = new_liste

                fig.add_shape(type="rect",
                                      x0=anomaly[0], y0=min(df_cells['Avg Num HSDPA Users']), x1=anomaly[1], y1=max(df_cells['Avg Num HSDPA Users']),
                                      line=dict(
                                          color="Crimson",
                                          width=2,
                                      ),
                                      )
                return (dcc.Graph(figure=fig))

            else:
                return (None)
        except:
            return(None)


@app.callback(Output('kpi_TotalDataTrafficDl&UL(GBytes)', 'children'),
 			[Input('graph_button_4', 'n_clicks'),
             Input('dropdown_cell', 'children'),
             Input('radio_pim', 'value')])
def kpi_tdt(n_clicks, cell, choice):
    
    if n_clicks >= 1 :

        try:
            cell = cell['props']['value']
            df = pd.read_csv('data_dash/df_clean.csv')

            table_cells = pd.read_csv('data_dash/table_cells.csv')
            anomaly = max(table_cells[table_cells.Cell_dropdown == cell]['Anomaly_range'])
            cell = max(table_cells[table_cells.Cell_dropdown == cell].Cell)

            if 'Total Data Traffic Dl&UL(GBytes)' in df.columns:

                new_liste = []
                liste = cell.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' ', '')
                    new = new.replace("'", '')
                    new_liste.append(new)
                cells = new_liste
                df_cells = pd.DataFrame()
                for cell in cells:
                    df_cell = df[df.Cell == cell].sort_index()
                    df_cells = pd.concat([df_cells, df_cell])

                if choice == 'PIM':
                    fig = px.line(df_cells, x='Time', y='Total Data Traffic Dl&UL(GBytes)')
                else:
                    fig = px.line(df_cells, x='Time', y='Total Data Traffic Dl&UL(GBytes)', color='Cell')

                new_liste = []
                liste = anomaly.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' Timestamp(', '')
                    new = new.replace('Timestamp(', '')
                    new = new.replace(')', '')
                    new = new.replace("'", '')
                    new_liste.append(pd.to_datetime(new))
                anomaly = new_liste

                fig.add_shape(type="rect",
                                      x0=anomaly[0], y0=min(df_cells['Total Data Traffic Dl&UL(GBytes)']), x1=anomaly[1], y1=max(df_cells['Total Data Traffic Dl&UL(GBytes)']),
                                      line=dict(
                                          color="Crimson",
                                          width=2,
                                      ),
                                      )

                return (dcc.Graph(figure=fig))

            else:
                return (None)
        except:
            return(None)


@app.callback(Output('kpi_3GTRAFFICSPEECH', 'children'),
 			[Input('graph_button_4', 'n_clicks'),
             Input('dropdown_cell', 'children'),
             Input('radio_pim', 'value')])
def kpi_trafficspeech(n_clicks, cell, choice):
    
    if n_clicks >= 1 :

        try:
            cell = cell['props']['value']
            df = pd.read_csv('data_dash/df_clean.csv')

            table_cells = pd.read_csv('data_dash/table_cells.csv')
            anomaly = max(table_cells[table_cells.Cell_dropdown == cell]['Anomaly_range'])
            cell = max(table_cells[table_cells.Cell_dropdown == cell].Cell)

            if '3G TRAFFIC SPEECH' in df.columns:

                new_liste = []
                liste = cell.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' ', '')
                    new = new.replace("'", '')
                    new_liste.append(new)
                cells = new_liste
                df_cells = pd.DataFrame()
                for cell in cells:
                    df_cell = df[df.Cell == cell].sort_index()
                    df_cells = pd.concat([df_cells, df_cell])

                if choice == 'PIM':
                    fig = px.line(df_cells, x='Time', y='3G TRAFFIC SPEECH')
                else:
                    fig = px.line(df_cells, x='Time', y='3G TRAFFIC SPEECH', color='Cell')

                new_liste = []
                liste = anomaly.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' Timestamp(', '')
                    new = new.replace('Timestamp(', '')
                    new = new.replace(')', '')
                    new = new.replace("'", '')
                    new_liste.append(pd.to_datetime(new))
                anomaly = new_liste

                fig.add_shape(type="rect",
                                      x0=anomaly[0], y0=min(df_cells['3G TRAFFIC SPEECH']), x1=anomaly[1], y1=max(df_cells['3G TRAFFIC SPEECH']),
                                      line=dict(
                                          color="Crimson",
                                          width=2,
                                      ),
                                      )

                return (dcc.Graph(figure=fig))

            else:
                return (None)
        except:
            return(None)
    
@app.callback(Output('kpi_3G&3G+ULTraffic', 'children'),
 			[Input('graph_button_4', 'n_clicks'),
             Input('dropdown_cell', 'children'),
             Input('radio_pim', 'value')])
def kpi_ultraff(n_clicks, cell, choice):

    if n_clicks >= 1 :

        try:
            cell = cell['props']['value']
            df = pd.read_csv('data_dash/df_clean.csv')

            table_cells = pd.read_csv('data_dash/table_cells.csv')
            anomaly = max(table_cells[table_cells.Cell_dropdown == cell]['Anomaly_range'])
            cell = max(table_cells[table_cells.Cell_dropdown == cell].Cell)

            if '3G&3G+ UL Traffic' in df.columns:

                new_liste = []
                liste = cell.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' ', '')
                    new = new.replace("'", '')
                    new_liste.append(new)
                cells = new_liste
                df_cells = pd.DataFrame()
                for cell in cells:
                    df_cell = df[df.Cell == cell].sort_index()
                    df_cells = pd.concat([df_cells, df_cell])

                if choice == 'PIM':
                    fig = px.line(df_cells, x='Time', y='3G&3G+ UL Traffic')
                else:
                    fig = px.line(df_cells, x='Time', y='3G&3G+ UL Traffic', color='Cell')

                new_liste = []
                liste = anomaly.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' Timestamp(', '')
                    new = new.replace('Timestamp(', '')
                    new = new.replace(')', '')
                    new = new.replace("'", '')
                    new_liste.append(pd.to_datetime(new))
                anomaly = new_liste

                fig.add_shape(type="rect",
                                      x0=anomaly[0], y0=min(df_cells['3G&3G+ UL Traffic']), x1=anomaly[1], y1=max(df_cells['3G&3G+ UL Traffic']),
                                      line=dict(
                                          color="Crimson",
                                          width=2,
                                      ),
                                      )

                return (dcc.Graph(figure=fig))

            else:
                return (None)
        except:
            return(None)


@app.callback(Output('kpi_dcrps', 'children'),
 			[Input('graph_button_4', 'n_clicks'),
             Input('dropdown_cell', 'children'),
             Input('radio_pim', 'value')])
def kpi_dcrps(n_clicks, cell, choice):
    
    if n_clicks >= 1 :

        try:
            cell = cell['props']['value']
            df = pd.read_csv('data_dash/df_clean.csv')

            table_cells = pd.read_csv('data_dash/table_cells.csv')
            anomaly = max(table_cells[table_cells.Cell_dropdown == cell]['Anomaly_range'])
            cell = max(table_cells[table_cells.Cell_dropdown == cell].Cell)

            if 'DCR-PS' in df.columns:

                new_liste = []
                liste = cell.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' ', '')
                    new = new.replace("'", '')
                    new_liste.append(new)
                cells = new_liste
                df_cells = pd.DataFrame()
                for cell in cells:
                    df_cell = df[df.Cell == cell].sort_index()
                    df_cells = pd.concat([df_cells, df_cell])

                if choice == 'PIM':
                    fig = px.line(df_cells, x='Time', y='DCR-PS')
                else:
                    fig = px.line(df_cells, x='Time', y='DCR-PS', color='Cell')

                new_liste = []
                liste = anomaly.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' Timestamp(', '')
                    new = new.replace('Timestamp(', '')
                    new = new.replace(')', '')
                    new = new.replace("'", '')
                    new_liste.append(pd.to_datetime(new))
                anomaly = new_liste

                fig.add_shape(type="rect",
                                      x0=anomaly[0], y0=min(df_cells['DCR-PS']), x1=anomaly[1], y1=max(df_cells['DCR-PS']),
                                      line=dict(
                                          color="Crimson",
                                          width=2,
                                      ),
                                      )

                return (dcc.Graph(figure=fig))

            else:
                return (None)
        except:
            return(None)


@app.callback(Output('kpi_cssrcs', 'children'),
 			[Input('graph_button_4', 'n_clicks'),
             Input('dropdown_cell', 'children'),
             Input('radio_pim', 'value')])
def kpi_cssrcs(n_clicks, cell, choice):
    
    if n_clicks >= 1 :

        try:
            cell = cell['props']['value']
            df = pd.read_csv('data_dash/df_clean.csv')

            table_cells = pd.read_csv('data_dash/table_cells.csv')
            anomaly = max(table_cells[table_cells.Cell_dropdown == cell]['Anomaly_range'])
            cell = max(table_cells[table_cells.Cell_dropdown == cell].Cell)

            if 'CSSR-CS' in df.columns:

                new_liste = []
                liste = cell.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' ', '')
                    new = new.replace("'", '')
                    new_liste.append(new)
                cells = new_liste
                df_cells = pd.DataFrame()
                for cell in cells:
                    df_cell = df[df.Cell == cell].sort_index()
                    df_cells = pd.concat([df_cells, df_cell])

                if choice == 'PIM':
                    fig = px.line(df_cells, x='Time', y='CSSR-CS')
                else:
                    fig = px.line(df_cells, x='Time', y='CSSR-CS', color='Cell')

                new_liste = []
                liste = anomaly.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' Timestamp(', '')
                    new = new.replace('Timestamp(', '')
                    new = new.replace(')', '')
                    new = new.replace("'", '')
                    new_liste.append(pd.to_datetime(new))
                anomaly = new_liste

                fig.add_shape(type="rect",
                                      x0=anomaly[0], y0=min(df_cells['CSSR-CS']), x1=anomaly[1], y1=max(df_cells['CSSR-CS']),
                                      line=dict(
                                          color="Crimson",
                                          width=2,
                                      ),
                                      )

                return (dcc.Graph(figure=fig))

            else:
                return (None)
        except:
            return(None)


@app.callback(Output('kpi_cssrps', 'children'),
 			[Input('graph_button_4', 'n_clicks'),
             Input('dropdown_cell', 'children'),
             Input('radio_pim', 'value')])
def kpi_cssrps(n_clicks, cell, choice):
    
    if n_clicks >= 1 :

        try:
            cell = cell['props']['value']
            df = pd.read_csv('data_dash/df_clean.csv')

            table_cells = pd.read_csv('data_dash/table_cells.csv')
            anomaly = max(table_cells[table_cells.Cell_dropdown == cell]['Anomaly_range'])
            cell = max(table_cells[table_cells.Cell_dropdown == cell].Cell)

            if 'CSSR-PS' in df.columns:

                new_liste = []
                liste = cell.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' ', '')
                    new = new.replace("'", '')
                    new_liste.append(new)
                cells = new_liste
                df_cells = pd.DataFrame()
                for cell in cells:
                    df_cell = df[df.Cell == cell].sort_index()
                    df_cells = pd.concat([df_cells, df_cell])

                if choice == 'PIM':
                    fig = px.line(df_cells, x='Time', y='CSSR-PS')
                else:
                    fig = px.line(df_cells, x='Time', y='CSSR-PS', color='Cell')

                new_liste = []
                liste = anomaly.split(',')
                for new in liste:
                    new = new.replace('[', '')
                    new = new.replace(']', '')
                    new = new.replace(' Timestamp(', '')
                    new = new.replace('Timestamp(', '')
                    new = new.replace(')', '')
                    new = new.replace("'", '')
                    new_liste.append(pd.to_datetime(new))
                anomaly = new_liste

                fig.add_shape(type="rect",
                                      x0=anomaly[0], y0=min(df_cells['CSSR-PS']), x1=anomaly[1], y1=max(df_cells['CSSR-PS']),
                                      line=dict(
                                          color="Crimson",
                                          width=2,
                                      ),
                                      )

                return (dcc.Graph(figure=fig))

            else:
                return (None)
        except:
            return(None)

def parse_contents(contents, filename, date):
    content_type, content_string = contents.split(',')

    decoded = base64.b64decode(content_string)
    try:
        if 'csv' in filename:
            # Assume that the user uploaded a CSV file
            df = pd.read_csv(
                io.StringIO(decoded.decode('utf-8')))
        elif 'xls' in filename:
            # Assume that the user uploaded an excel file
            df = pd.read_excel(io.BytesIO(decoded))

        if 'Latitude' in df.columns :
            df.to_csv('data_dash/df_loc.csv')
        if 'RTWP' in df.columns :
            df.to_csv('data_dash/df_total.csv')
        if 'Crown' in df.columns :
            df.to_csv('data_dash/crown_data.csv')

    except Exception as e:
        print(e)
        return html.Div([
            'There was an error processing this file.'
        ])

    return html.Div([
        html.H5(filename),
        html.H6(datetime.datetime.fromtimestamp(date)),

        dash_table.DataTable(
            data=df.to_dict('records'),
            columns=[{'name': i, 'id': i} for i in df.columns]
        ),

        html.Hr(),  # horizontal line

        # For debugging, display the raw contents provided by the web browser
        html.Div('Raw Content'),
        html.Pre(contents[0:200] + '...', style={
            'whiteSpace': 'pre-wrap',
            'wordBreak': 'break-all'
        })
    ])

@app.callback(Output('output-data-upload', 'children'),
              Input('upload-data', 'contents'),
              State('upload-data', 'filename'),
              State('upload-data', 'last_modified'))
def update_output(list_of_contents, list_of_names, list_of_dates):
    for filename in os.listdir('data_dash'):
        if filename == 'df_total.csv' :
            os.remove("data_dash/" + filename)

    if list_of_contents is not None:
        children = [
            parse_contents(c, n, d) for c, n, d in
            zip(list_of_contents, list_of_names, list_of_dates)]
        return html.H6('Data imported', style = {'margin-left':'180px'})

@app.callback(Output('output-data_loc-upload', 'children'),
              Input('upload-data_loc', 'contents'),
              State('upload-data_loc', 'filename'),
              State('upload-data_loc', 'last_modified'))
def update_output(list_of_contents, list_of_names, list_of_dates):
    for filename in os.listdir('data_dash'):
        print(filename)
        if filename == 'df_loc.csv' :
            os.remove("data_dash/" + filename)

    if list_of_contents is not None:
        children = [
            parse_contents(c, n, d) for c, n, d in
            zip(list_of_contents, list_of_names, list_of_dates)]
        return html.H6('Data imported', style = {'margin-left':'180px'})

@app.callback(Output('output-data_crown-upload', 'children'),
              Input('upload-data_crown', 'contents'),
              State('upload-data_crown', 'filename'),
              State('upload-data_crown', 'last_modified'))
def update_output(list_of_contents, list_of_names, list_of_dates):
    for filename in os.listdir('data_dash'):
        print(filename)
        if filename == 'crown_data.csv' :
            os.remove("data_dash/" + filename)

    if list_of_contents is not None:
        children = [
            parse_contents(c, n, d) for c, n, d in
            zip(list_of_contents, list_of_names, list_of_dates)]
        return html.H6('Data imported', style = {'margin-left':'180px'})

@app.callback(Output('export_excel_text', 'children'),
              Input('excel_button', 'n_clicks'))
def export_excel(n_clicks):
    def _gutter(idx, offset, max_val):
        """
        When deleting rows and columns are deleted we rely on overwriting.
        This may not be the case for a large offset on small set of cells:
        range(cells_to_delete) > range(cell_to_be_moved)
        """
        gutter = range(max(max_val + 1 - offset, idx), min(idx + offset, max_val) + 1)
        return gutter

    ####recoded openpyxl methods because of old versions installed on edennet 19

    def move_cell(sheet, row, column, row_offset, col_offset, translate=False):
        """
        Move a cell from one place to another.
        Delete at old index
        Rebase coordinate
        """
        cell = sheet._get_cell(row, column)
        new_row = cell.row + row_offset
        new_col = cell.column + col_offset
        sheet._cells[new_row, new_col] = cell
        del sheet._cells[(cell.row, cell.column)]
        cell.row = new_row
        cell.column = new_col
        if translate and cell.data_type == "f":
            t = Translator(cell.value, cell.coordinate)
            cell.value = t.translate_formula(row_delta=row_offset, col_delta=col_offset)

        return sheet

    def move_cells(sheet, min_row=None, min_col=None, offset=0, row_or_col="row"):
        """
        Move either rows or columns around by the offset
        """
        reverse = offset > 0  # start at the end if inserting
        row_offset = 0
        col_offset = 0

        # need to make affected ranges contiguous
        if row_or_col == 'row':
            cells = sheet.iter_rows(min_row=min_row)
            row_offset = offset
            key = 0
        else:
            cells = sheet.iter_cols(min_col=min_col)
            col_offset = offset
            key = 1
        cells = list(cells)

        for row, column in sorted(sheet._cells, key=itemgetter(key), reverse=reverse):
            if min_row and row < min_row:
                continue
            elif min_col and column < min_col:
                continue

            sheet._move_cell(row, column, row_offset, col_offset)

        return sheet

    def delete_rows(sheet, idx, amount=1):

        remainder = _gutter(idx, amount, sheet.max_row)

        # sheet._move_cells(min_row=idx + amount, offset=-amount, row_or_col="row")
        sheet = move_cells(sheet, min_row=idx + amount, offset=-amount, row_or_col="row")

        # calculating min and max col is an expensive operation, do it only once
        min_col = sheet.min_column
        max_col = sheet.max_column + 1
        for row in remainder:
            for col in range(min_col, max_col):
                if (row, col) in sheet._cells:
                    del sheet._cells[row, col]
        sheet._current_row = sheet.max_row
        if not sheet._cells:
            sheet._current_row = 0

        return sheet

    if n_clicks >= 1 :
        main_output = pd.read_csv('data_dash/main_output.csv')
        main_output = main_output[['Cell','last_anomaly','last_anomaly_range']]
        detailled_output = pd.read_csv('data_dash/detailled_output.csv')
        del detailled_output['Unnamed: 0']
        miss_data = pd.read_csv('data_dash/missing_data.csv')
        df_data = pd.read_csv('data_dash/df_total.csv')
        external_interferences = pd.read_csv('data_dash/ext_interf.csv')
        PIM = pd.read_csv('data_dash/df_no_final.csv')
        anom = pd.read_csv('data_dash/anom.csv')

        df_main_output = main_output.values.tolist()
        del detailled_output['preclust']
        del detailled_output['last_anomaly']
        del detailled_output['last_anomaly_range']
        del detailled_output['final_ranking']
        df_detailled_output = detailled_output.values.tolist()
        external_interferences = external_interferences[['Neighbor','Since_anom_end','Anom_duration','qos_degradation','max_RTWP','Localization']]
        df_external_issues = external_interferences.values.tolist()
        PIM = PIM[['Cell','Since_anom_end','Anom_duration','qos_degradation','max_RTWP']]
        df_pim_issues = PIM.values.tolist()

        df_miss_data = [str(i) for i in miss_data['Cell']]

        # rtwp_data
        rtwp_data = pd.DataFrame()
        anom = anom.sort_values(by='final_ranking')
        for cell in anom.Cell.unique():
            anom_cell = anom[anom.Cell == cell]
            ranking = max(anom_cell.final_ranking)
            if len(str(max(anom_cell.final_ranking))) > 1:
                ranking = str(max(anom_cell.final_ranking))
            else:
                ranking = '0' + str(max(anom_cell.final_ranking))
            if cell not in df_miss_data:
                df_cell = df_data[df_data.Cell == cell].sort_values(by='Time')
                df_cell['Cell'] = str(ranking) + '_' + max(df_cell.Cell)
                df_cell = df_cell[['Cell', 'Time', 'RTWP']]
                df_cell['Time'] = [str(i) for i in df_cell['Time']]
                rtwp_data = pd.concat([rtwp_data, df_cell])
        rtwp_data['RTWP'] = [round(j, 3) for j in rtwp_data['RTWP']]

        df_rtwp_data = rtwp_data.values.tolist()

        wb = openpyxl.load_workbook('Output_Excel.xlsx')

        sheet1 = wb['main_output']
        sheet1 = delete_rows(sheet1, 3, sheet1.max_row + 1)

        for row in range(3, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                cell = sheet1.cell(row=row, column=col)
                if cell.value != None:
                    cell.value = None

        currentRow = 2
        for row in df_main_output:
            currentRow += 1
            currentCol = 1
            sheet1.cell(row=currentRow, column=currentCol).value = row[0]
            currentCol += 1
            sheet1.cell(row=currentRow, column=currentCol).value = row[1]
            currentCol += 1
            sheet1.cell(row=currentRow, column=currentCol).value = row[2]

        sheet1 = wb['detailled_output']
        sheet1 = delete_rows(sheet1, 3, sheet1.max_row + 1)

        currentRow = 2
        for row in df_detailled_output:
            currentRow += 1
            currentCol = 0
            for val in row:
                currentCol += 1
                sheet1.cell(row=currentRow, column=currentCol).value = val

        sheet1 = wb['external_interferences']
        sheet1 = delete_rows(sheet1, 3, sheet1.max_row + 1)

        for row in range(3, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                cell = sheet1.cell(row=row, column=col)
                if cell.value != None:
                    cell.value = None

        currentRow = 2
        for row in df_external_issues:
            currentRow += 1
            currentCol = 0
            for val in row:
                currentCol += 1
                sheet1.cell(row=currentRow, column=currentCol).value = str(val)

        sheet1 = wb['PIM']
        sheet1 = delete_rows(sheet1, 3, sheet1.max_row + 1)

        for row in range(3, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                cell = sheet1.cell(row=row, column=col)
                if cell.value != None:
                    cell.value = None

        currentRow = 2
        for row in df_pim_issues:
            currentRow += 1
            currentCol = 0
            for val in row:
                currentCol += 1
                sheet1.cell(row=currentRow, column=currentCol).value = str(val)

        sheet1 = wb['missing_data']
        sheet1 = delete_rows(sheet1, 2, sheet1.max_row + 1)

        currentRow = 1
        currentCol = 1
        for row in df_miss_data:
            val = row
            currentRow += 1
            sheet1.cell(row=currentRow, column=currentCol).value = val

        sheet1 = wb['rtwp_data']

        max_row = sheet1.max_row + 1
        actual_len = len(df_rtwp_data)

        currentRow = 1
        for row in df_rtwp_data:
            currentRow += 1
            currentCol = 1
            sheet1.cell(row=currentRow, column=currentCol).value = row[0]
            currentCol += 1
            sheet1.cell(row=currentRow, column=currentCol).value = row[1]
            currentCol += 1
            sheet1.cell(row=currentRow, column=currentCol).value = row[2]

        if actual_len < max_row:
            for row in range(actual_len, max_row + 2):
                for col in [1, 2, 3, 4]:
                    cell = sheet1.cell(row=row, column=col)
                    if cell.value != None:
                        cell.value = None

        wb.save('Output_Excel.xlsx')

        return html.H6('Excel export done.', style = {'margin-left':'180px'})


#################### Lancement de l'application ####################

if __name__ == '__main__':
    # Déploiement en privé
    app.run_server(debug=True)

    # Déploiement en public
    #app.run_server(debug=False, host='0.0.0.0', port='8050')